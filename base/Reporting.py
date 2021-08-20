import pandas as pd
import numpy as np

from scipy.stats import chi2_contingency

from .model_refining import RefineBSPredictions
from .model_building import make_predictions

import matplotlib.pyplot as plt
import matplotlib.lines as mlines

from openpyxl import load_workbook
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Side, Border
from openpyxl.utils.dataframe import dataframe_to_rows

from statsmodels.stats.outliers_influence import variance_inflation_factor
from statsmodels.tools.tools import add_constant
import os

class Reporting:
    def __init__(self,training_set,testing_set,pdv_set,oot_set,bin_info,fine_tuneing=False):                
        self.training_set = training_set
        self.testing_set = testing_set
        self.validation_set = pdv_set
        self.bin_info = bin_info
        self.ootd_set = oot_set                
        self.feature_bin_mapping = {}
        self.datasets_info = {}
        self.final_reporting_data = pd.DataFrame()
        #self.initialize_data()
        self.pdv_isset = False
        self.oot_isset = False 
        self.results = {}  
        self.fine_tuneing = fine_tuneing

    def set_mlflow_metrics(self, result_dict):
        self.results = result_dict

    def get_mlflow_metrics(self):
        return self.results

    def generate_reports(self):
        #Start
        #TODO Change data
        temp_obj  = pd.read_pickle(os.getcwd()+'/base/temp/'+'model_building.pkl')   #TODO Look to replace
        #checks on oot & pdv TODO
        #checks on preprocessor.pkl TODO
        model_obj = pd.read_pickle(os.getcwd()+'/base/temp/'+'model_preprocessor.pkl')  #TODO Look to replace
        test_data = self.testing_set
        #pd.read_csv('temp/test-binned-result.csv')#Comment
        oot_data = self.ootd_set
        #pd.read_csv('temp/oot-binned-result.csv')#Comment
        pdv_data = self.validation_set
        #test_data.copy()#Comment

        train_bins = model_obj['model_bins']

        oot_set = False  #Comment
        pdv_set = False

        if not(oot_data is None):
            oot_set = True
        if not(pdv_data is None):
            pdv_set = True

        #if(oot_set):
        #    oot_data = oot_data.sample(frac=0.05)
        #    test_data = test_data.sample(frac=0.1)

        numeric_features = model_obj['numeric_features'].copy()
        categorical_features = model_obj['categorical_features'].copy()
        features = numeric_features.copy()
        features.extend(categorical_features)
        self.features = features
        
        #features = features.remove('yTarget') #Comment find a more reliable way
        pdv_df = pd.DataFrame()
        test_df = pd.DataFrame()
        oot_df = pd.DataFrame()
        
        #Runnind model on test data
        if(self.fine_tuneing):
            test_predictions =  RefineBSPredictions(test_data, model_obj)
        else:
            test_predictions = make_predictions(test_data, model_obj)
            #added model risk bands
        
        test_preds_bins = test_predictions['binned_variables']
        psi_train_df =  model_obj['model_risk_df_train']
        psi_test_df = test_predictions['model_risk_df'] #Risk Bands
        print('train --- ')
        print(psi_train_df.head(5))

        test_ypred = pd.Series(test_predictions['ypred'],name='predicted')
        
        test_df = pd.concat([test_preds_bins,test_ypred],axis=1)
        #test_df = test_df.sample(frac=0.2) #Comment Out
        #test_df.reset_index(inplace=True) #Comment Out
        test_df.to_csv(os.getcwd()+'/base/'+'temp/test-binned-result.csv',index=False) #Comment Out
        #TODO add Risk Band details here

        reporting_df_test = process_data(test_df,'Test',features)


        #checks on oot & pdv TODO
        #3 Getting ootd & making preds 
        if(oot_set):
            if(self.fine_tuneing):
                oot_predicted = RefineBSPredictions(oot_data,model_obj)
            else:
                oot_predicted = make_predictions(oot_data,model_obj)            
            psi_oot_df = oot_predicted['model_risk_df'] #Risk Bands
            print('oot --- ')
            print(psi_oot_df.head(5))
            oot_preds_bins = oot_predicted['binned_variables']
            #3.2 combining & write oot
            oot_ypred = pd.Series(oot_predicted['ypred'],name='predicted')
            oot_df = pd.concat([oot_preds_bins,oot_ypred],axis=1)
            oot_df.to_csv(os.getcwd()+'/base/'+'temp/oot-binned-result.csv',index=False)  #Comment
            #TODO add Risk Band details here
            #3.3 Creating reporting_df
            reporting_df_oot = process_data(oot_df,'OOT',features)

        #Comment
        
        #checks on oot & pdv TODO
        #3 Getting pdv & making preds 
        if(pdv_set):
            if(self.fine_tuneing):
                pdv_predicted = RefineBSPredictions(pdv_data,model_obj)
            else:
                pdv_predicted = make_predictions(pdv_data,model_obj)
                
            pdv_preds_bins = pdv_predicted['binned_variables']
            psi_pdv_df = pdv_predicted['model_risk_df'] #Risk Bands
            #3.2 combining & write oot
            pdv_ypred = pd.Series(pdv_preds_bins['ypred'],name='predicted')
            pdv_df = pd.concat([pdv_preds_bins,pdv_ypred],axis=1)
            pdv_df.to_csv(os.getcwd()+'/base/'+'temp/oot-binned-result.csv',index=False)  #Comment
            #TODO add Risk Band details here
            #3.3 Creating reporting_df
            reporting_df_pdv = process_data(oot_df,'PDV',features)



        #checks on oot & pdv TODO
        if(pdv_set and oot_set):
            reporting_df_final = pd.concat([reporting_df_test,reporting_df_oot,reporting_df_pdv],axis=0)
            model_band_psi = compute_model_band_psi(psi_train_df,psi_test_df,psi_pdv_df,psi_oot_df) #Risk Bands
        elif(pdv_set):
            reporting_df_final = pd.concat([reporting_df_test,reporting_df_pdv],axis=0)
            model_band_psi = compute_model_band_psi(psi_train_df,psi_test_df,psi_pdv_df,None) #Risk Bands
        else:
            reporting_df_final = pd.concat([reporting_df_test,reporting_df_oot],axis=0)
            model_band_psi = compute_model_band_psi(psi_train_df,psi_test_df,None,psi_oot_df) #Risk Bands

        
        reporting_df_final.to_csv(os.getcwd()+'/base/'+'Reports/Reporting_Data.csv',index=False)

        #-----------------     Begin Reporting        ----------------------------
        #Creating empty Reporting file 
        wb = openpyxl.Workbook() 
        ws = wb.active
        ws.title = '4.Bivariates'
        ws.sheet_view.showGridLines = False
        wb.save(filename=os.getcwd()+'/base/'+'Reports/Post_Train_Model_Report.xlsx')

        reporting_df_final = pd.read_csv(os.getcwd()+'/base/'+'Reports/Reporting_Data.csv')  #Comment
        features = reporting_df_final['Feature'].unique().tolist()


        #pkl_filename = "temp/oot_preds.pkl"

        #with open(pkl_filename, 'rb') as file: #Comment - Quick run 
        #    oot_preds_bins =  pickle.load(file)

        #pkl_filename = "temp/test_preds.pkl"    #Comment - Quick run 
        #with open(pkl_filename, 'rb') as file:
        #    test_preds_bins =  pickle.load(file)

        #pdv_preds_bins = test_preds_bins.copy() #Comment - Quick run 



        accuracy_dict = {}

        #TODO Consider moving into a sepearate function

        ks_df  = pd.DataFrame({'Actuals':test_predictions['ytest']})
        y_pred = pd.Series(test_predictions['y_pred_probability'], name = 'Probability')#TODO will not work with Logit & sklearn seamlessly. Find workaround
        ks_df  = pd.concat([ks_df,y_pred], axis=1)

        accuracy_dict['AUC_Test'] = round(test_predictions['roc_auc_score'],5)
        accuracy_dict['KS_Test'] = round(compute_ks(ks_df,'Probability','Actuals'),5)

        if(oot_set):
            ks_df  = pd.DataFrame({'Actuals':oot_predicted['ytest']})
            y_pred = pd.Series(oot_predicted['y_pred_probability'], name = 'Probability')#TODO will not work with Logit & sklearn seamlessly. Find workaround
            ks_df  = pd.concat([ks_df,y_pred], axis=1)

            accuracy_dict['AUC_OOT'] = round(oot_predicted['roc_auc_score'],5)
            accuracy_dict['KS_OOT'] = round(compute_ks(ks_df,'Probability','Actuals'),5)

        if(pdv_set):
            ks_df  = pd.DataFrame({'Actuals':pdv_predicted['ytest']})
            y_pred = pd.Series(pdv_predicted['y_pred_probability'], name = 'Probability')#TODO will not work with Logit & sklearn seamlessly. Find workaround
            ks_df  = pd.concat([ks_df,y_pred], axis=1)

            accuracy_dict['AUC_PDV'] = round(pdv_predicted['roc_auc_score'],5)
            accuracy_dict['KS_PDV'] = round(compute_ks(ks_df,'Probability','Actuals'),5)

        accuracy_df =  pd.DataFrame(accuracy_dict,index=[0])
                
        if(pdv_set & oot_set):
            build_model_report(model_obj,test_predictions,oot_predicted,pdv_predicted)
        elif(pdv_set):
            build_model_report(model_obj,test_predictions,None,pdv_predicted)
        else:
            build_model_report(model_obj,test_predictions,oot_predicted,None)
        append_model_report(accuracy_df,'2.AUC_KS')
        append_model_report(reporting_df_final,"3.ReportingData")
        plot_all_bads(reporting_df_final,features)
        build_iv_report(reporting_df_final,features)
        build_cramersv_report(test_df,oot_df,pdv_df,features,"predicted") #TODO find a way to not depend on this key
        build_psi_report(reporting_df_final,features)
        append_model_report(model_band_psi,"9. ModelBand_PSI_Summary") 
        plot_gains_lift(model_band_psi)
        model_band_psi_dict = {}
        model_band_psi_dict['PSI_Train_Test'] = [model_band_psi['PSI_Test'].sum()]
        if oot_set:
            model_band_psi_dict['PSI_Train_OOT'] = [model_band_psi['PSI_OOT'].sum()]
        if pdv_set:
            model_band_psi_dict['PSI_Train_PDV'] = [model_band_psi['PSI_PDV'].sum()]
        append_model_report(pd.DataFrame(model_band_psi_dict),"10. ModelBand_PSI_Final") 

        wb = openpyxl.load_workbook(os.getcwd()+'/base/'+'Reports/Post_Train_Model_Report.xlsx')
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for col in map(chr, range(ord('A'), ord('Z')+1)):
                ws.column_dimensions[col].width = 11.5

        wb.save(filename=os.getcwd()+'/base/'+'Reports/Post_Train_Model_Report.xlsx')
        self.set_mlflow_metrics(accuracy_dict)

def append_model_report(df,name):
    book = load_workbook(os.getcwd()+'/base/'+'Reports/Post_Train_Model_Report.xlsx')
    writer = pd.ExcelWriter(os.getcwd()+'/base/'+'Reports/Post_Train_Model_Report.xlsx',engine='openpyxl')
    writer.book = book    
    df.to_excel(writer,sheet_name=name,index=False)
    book._sheets.sort(key=lambda ws: ws.title)
    writer.save()
    writer.close()

def getCramerv(c,n,contigency):
        phi2 = c/n
        r,k = contigency.shape
        phi2corr = max(0.0, phi2 - (((k-1)*(r-1))/(n-1)))    
        rcorr = r - ((r-1)**2)/(n-1)
        kcorr = k - ((k-1)**2)/(n-1)
        if( min( (kcorr-1), (rcorr-1)) == 0):
            return 0
        cramers_v =  round(np.sqrt(phi2corr / min( (kcorr-1), (rcorr-1))),4) 
        return cramers_v
    
def getChiSquare(data, features,dependent_feature,source):
        chi_sqaure_data = {'Feature':[],'Chi-Square':[],'P-Value':[],'CramersV':[]}        
        for feature in features:
            if(feature == dependent_feature):
                continue
            else:            
                contigency= pd.crosstab(data[feature],data[dependent_feature])
                c, p, dof, expected = chi2_contingency(contigency)
                n =  sum(contigency.sum())
                chi_sqaure_data["Feature"].append(feature)
                chi_sqaure_data["Chi-Square"].append(round(c,4))
                chi_sqaure_data["P-Value"].append(round(p,4))
                chi_sqaure_data['CramersV'].append(getCramerv(c,n,contigency))
        chi_sqaure_pd = pd.DataFrame(chi_sqaure_data,columns=['Feature','Chi-Square','P-Value','CramersV'])
        chi_sqaure_pd['Source'] = source
        return chi_sqaure_pd

def compute_vif(dataset,features):        
    feature_list = []
    for feature in features:
        if(dataset.dtypes[feature] not in ['int64','float64']):
            continue
        else:
            feature_list.append(feature)    
    vif_ds = add_constant(dataset[feature_list])
    vifs = pd.Series([variance_inflation_factor(vif_ds.values, i) for i in range(vif_ds.shape[1])], index=vif_ds.columns)
    return vifs

def process_data(cons_df,source,features):
    consolidated_feature_count_df = pd.DataFrame()
    for feature in features:    
        if (feature == 'predicted'):
            continue
        else:
            #List of Bins of a var with pred value
            temp_df = cons_df[[feature,'predicted']]
            new_df = temp_df.groupby([feature]).sum()
            new_df['Total'] = temp_df.groupby([feature]).count() 
            new_df.reset_index(inplace=True)
            new_df.columns = ['Bins','Goods','Total']
            new_df['Bads'] = new_df['Total'] - new_df['Goods']
            # - temp_df.groupby([feature]).sum()
            new_df['Feature'] = feature
            new_df = new_df[['Feature','Bins','Total','Goods','Bads']]
            total_count = new_df[new_df['Feature']==feature]['Total'].sum()
            #print(total_count)
            #print("Total = "+str(total_count))
            good_count = new_df[new_df['Feature']==feature]['Goods'].sum()
            #print("Good = "+str(good_count))
            bad_count = new_df[new_df['Feature']==feature]['Bads'].sum()
            #print("Bad = "+str(bad_count))
            #new_df['Bad Rate'] = round((new_df['Total']/total_count),2)
            new_df['Pop_Percent'] = round(new_df['Total'] / total_count,4)
            new_df['Good_Rate'] = round(new_df['Goods'] / new_df['Total'],4) 
            new_df['Bad_Rate'] = round(new_df['Bads'] / new_df['Total'],4) 
            new_df['Good_Percent'] = round(new_df['Goods'] / good_count,4) 
            new_df['Bad_Percent'] = round(new_df['Bads'] / bad_count,4) 
            new_df['WOE'] = round(np.log(new_df['Good_Percent']/new_df['Bad_Percent']),4)
            new_df['IV'] = ((new_df['Goods']/good_count) - (new_df['Bads']/bad_count)) * new_df['WOE']
            new_df['WOE'] = round(new_df['WOE'].replace(to_replace=[np.inf,-np.inf],value=[0.00,0.00]),4)
            new_df['IV'] = round(new_df['IV'].replace(to_replace=[np.inf,-np.inf],value=[0.00,0.00]),4)

            # new_df['Cumulative_Goods'] = round(new_df['Good_Percent'].cumsum(),4)
            # new_df['Cumulative_Bads'] = round(new_df['Good_Percent'].cumsum(),4)
            # new_df['KS'] = abs(new_df['Cumulative_Goods'] - new_df['Cumulative_Bads'])

            #new_df['Population'] = (new_df['Goods'] - new_df['Bads']) * new_df['WOE']
        consolidated_feature_count_df = pd.concat([consolidated_feature_count_df,new_df],axis=0)
        #print(consolidated_feature_count_df)
    consolidated_feature_count_df['Source'] = source    
    #consolidated_feature_count_df.to_csv(os.getcwd()+'/base/'+'final.csv',ignore_index=True)
    return consolidated_feature_count_df

def plot_bads(bins, total_list_oot, badrate_list_oot,source,feature,location='temp'):
    fig, ax = plt.subplots()
    width = 0.3
    x = np.arange(len(bins))
    #bins = list(map((lambda x:  str(index(x)+1)+x),bins))
    for ind in range(len(bins)):        
        modified_bin = str(ind+1)+". "+str(bins[ind])
        bins[ind] = modified_bin
    #For dual bar charts
    #bad_rates = ax.bar(x - width/2, badrate_list_oot , width, label='% Bad Rate',color='orange')
    #For line graphs
    total_rates = ax.bar(x, total_list_oot , width, label='Population %',color='orange')
    #ax2 = ax.twinx()
    dual_axis = ax.twinx()
    dual_axis.plot(x, badrate_list_oot,color='blue', label='Bad Rate %')
    dual_axis.set_ylabel('Bad Rate %')

    for x1,y1 in zip(x,badrate_list_oot):
        dual_axis.annotate(str(badrate_list_oot[x1]), xy=(x1,y1))
    
    
    
    #rects2 = ax.bar(x + width/2 , total_list_oot ,  width,label='% Population')

    ax.set_ylabel('Percentage %')
    ax.set_xlabel(feature)

    ax.set_xticks(x[0:len(x):2])
    ax.set_xticklabels(bins[0:len(x):2])
    
    #ax.tick_params(axis='x', rotation=70)
    
    totals_line = mlines.Line2D([], [], color='orange',label='Population %' )
    bads_line = mlines.Line2D([], [], color='blue', label='Bad Rate %')
    plt.legend(handles=[bads_line,totals_line],loc='best')
    #ax.legend(loc='best')
    #dual_axis.legend(loc='best')

    ax.bar_label(total_rates,padding=3)
    #ax.bar_label(rects2, padding=3)

    #fig.tight_layout()
    #plt.show()
    #CHECK
    path = os.getcwd()+ '/base/temp/images/'+ feature + '-'+source+'.png'
    plt.savefig(path)
    plt.clf()

    wb = openpyxl.load_workbook(os.getcwd()+'/base/'+'Reports/Post_Train_Model_Report.xlsx')
    if not("4.Bivariates" in wb.sheetnames): 
        ws =  wb.create_sheet("Bivariates")
        ws.sheet_view.showGridLines = False
    else:
        ws = wb['4.Bivariates']

    img = openpyxl.drawing.image.Image(path)

    for cellno in range(2,1000,33):        
        col = 'A'
        if(source == 'oot'):
            col = 'K'
        elif(source == 'pdv'):
            col = 'U'
        #print(col+str(cellno))
        cell = ws[col+str(cellno)]        
        #print(cell)
        if cell.value is None:
            cell.value = str(feature+'-'+source)
            cell.font = Font(name='Calibri',size=14,bold=True)
            #print(cell.value)
            img_cell = col+str(cellno+3)
            #print(img_cell)
            #img.anchor(cell)
            ws.add_image(img, img_cell)            
            break          
    wb.save(os.getcwd()+'/base/'+'Reports/Post_Train_Model_Report.xlsx')

def plot_all_bads(final_df,features):
    for feature in features:
        bins = final_df[ (final_df['Source'] == 'OOT') & (final_df['Feature'] == feature) ]['Bins'].tolist()
        if(bins is None) or (len(bins) == 0):
            continue
        total_list = final_df[  (final_df['Source'] == 'OOT') & (final_df['Feature'] == feature)]['Pop_Percent'].tolist()
        badrate_list = final_df[ (final_df['Source'] == 'OOT') & (final_df['Feature'] == feature)]['Bad_Percent'].tolist()

        for i in range(len(total_list)):
            total_list[i] = round(total_list[i] * 100.0,2)
        for i in range(len(badrate_list)):
            badrate_list[i] = round(badrate_list[i] * 100.0,2)

        plot_bads(bins, total_list, badrate_list,source='oot',feature=feature,location='temp/img')

        bins = final_df[ (final_df['Source'] == 'Test') & (final_df['Feature'] == feature) ]['Bins'].tolist()
        if(bins is None) or (len(bins) == 0):
            continue
        total_list = final_df[  (final_df['Source'] == 'Test') & (final_df['Feature'] == feature)]['Pop_Percent'].tolist()
        badrate_list = final_df[ (final_df['Source'] == 'Test') & (final_df['Feature'] == feature)]['Bad_Percent'].tolist()
    
        for i in range(len(total_list)):
            total_list[i] = round(total_list[i] * 100.0,2)
        for i in range(len(badrate_list)):
            badrate_list[i] = round(badrate_list[i] * 100.0,2)


        plot_bads(bins, total_list, badrate_list,source='test',feature=feature,location=os.getcwd()+'/base/'+'temp/img')

        
        bins = final_df[ (final_df['Source'] == 'PDV') & (final_df['Feature'] == feature) ]['Bins'].tolist()
        if(bins is None) or (len(bins) == 0):            
            continue
        total_list = final_df[  (final_df['Source'] == 'PDV') & (final_df['Feature'] == feature)]['Pop_Percent'].tolist()
        badrate_list = final_df[ (final_df['Source'] == 'PDV') & (final_df['Feature'] == feature)]['Bad_Percent'].tolist()
    
        for i in range(len(total_list)):
            total_list[i] = round(total_list[i] * 100.0,2)
        for i in range(len(badrate_list)):
            badrate_list[i] = round(badrate_list[i] * 100.0,2)
        
        plot_bads(bins, total_list, badrate_list,source='pdv',feature=feature,location=os.getcwd()+'/base/'+'temp/img')

def build_psi_report(final_df,features):
    psi_final_df = pd.DataFrame()
    oot_set = False
    pdv_set = False
    #TODO check on oot & pdv sets
    for feature in features:   
        test_data = final_df[ (final_df['Source']=='Test') & (final_df['Feature']==feature) ]
        oot_data = final_df[ (final_df['Source']=='OOT') & (final_df['Feature']==feature) ]
        pdv_data = final_df[ (final_df['Source']=='PDV') & (final_df['Feature']==feature) ] 
        #final_df[ (final_df['Source']=='PDV') & (final_df['Feature']==feature) ]
        #print('pdv')
        #print(pdv_data.head(2))
        if( (oot_data is None) and (pdv_data is None)):
            print('No OOT & PDV data provided hence skipping PSI calculation')
            return
        
        if(not(oot_data is None) and not(oot_data.empty)):
            oot_set = True
        if(not(pdv_data is None) and not(pdv_data.empty)):
            pdv_set = True
        
        if(oot_set and pdv_set):
            combined_df = test_data[['Feature','Bins','Pop_Percent']].merge(oot_data[['Bins','Pop_Percent']],on='Bins',how='outer',suffixes=('_test', '_oot'))
            combined_df = combined_df.merge(pdv_data[['Bins','Pop_Percent']],on='Bins',how='outer')            
            combined_df.rename(columns = {'Pop_Percent':'Pop_Percent_pdv'},inplace=True)
        elif oot_set:
            combined_df = test_data[['Feature','Bins','Pop_Percent']].merge(oot_data[['Bins','Pop_Percent']],on='Bins',how='outer',suffixes=('_test', '_oot'))            
        else:
            combined_df = test_data[['Feature','Bins','Pop_Percent']].merge(pdv_data[['Bins','Pop_Percent']],on='Bins',how='outer',suffixes=('_test', '_pdv'))        
        psi_final_df = pd.concat([psi_final_df,combined_df])
    
    if(oot_set):
        psi_final_df['PSI_test_oot'] = round( (psi_final_df['Pop_Percent_test'] - psi_final_df['Pop_Percent_oot']) * np.log(psi_final_df['Pop_Percent_test']/psi_final_df['Pop_Percent_oot']),4)
        psi_final_df['PSI_test_oot'] = psi_final_df['PSI_test_oot'].replace(to_replace=[np.inf,-np.inf],value=0)
    if(pdv_set):
        psi_final_df['PSI_test_pdv'] = round((psi_final_df['Pop_Percent_test'] - psi_final_df['Pop_Percent_pdv']) * np.log(psi_final_df['Pop_Percent_test']/psi_final_df['Pop_Percent_pdv']),4)
        psi_final_df['PSI_test_pdv'] = psi_final_df['Pop_Percent_test'].replace(to_replace=[np.inf,-np.inf],value=0)

    summarized_psi = pd.DataFrame()    
    if(oot_set and pdv_set):
        oot = round(psi_final_df.groupby('Feature')['PSI_test_oot'].sum(),4) 
        #oot.rename('PSI_OOT')
        pdv = round(psi_final_df.groupby('Feature')['PSI_test_pdv'].sum(),4)
        #pdv.rename('PSI_PDV')
        temp = pdv.reset_index(name='PSI_PDV')
        #print(temp.dtypes)  
        summarized_psi = oot.reset_index(name='PSI_OOT')        
        #print(summarized_psi.dtypes)
        summarized_psi = summarized_psi.merge(temp,on='Feature',how='inner',suffixes=None)
        
    else:
        if(oot_set):
            psi_temp = round(psi_final_df.groupby('Feature')['PSI_test_oot'].sum(),4)
            psi_temp.rename('OOT_PSI')
            summarized_psi = pd.concat([summarized_psi,psi_temp])
            summarized_psi.reset_index(inplace=True)
            summarized_psi.rename(columns={"index":"Feature"})
        else:
            psi_temp = round(psi_final_df.groupby('Feature')['PSI_test_pdv'].sum(),4)
            psi_temp.rename('PDV_PSI')
            summarized_psi = pd.concat([summarized_psi,psi_temp])
            summarized_psi.reset_index(inplace=True)
            summarized_psi.rename(columns={"index":"Feature"})


    append_model_report(psi_final_df,'7.CSI_Detailed')
    append_model_report(summarized_psi,'8.CSI_Summary')
    
    #psi_final_df.to_csv(os.getcwd()+'/base/'+'temp/combined_bins.csv')        

def build_cramersv_report(test_data,oot_data,pdv_data,features,dependent_feature):

    oot_set = False
    pdv_set = False
    feature_list = []
    for feature in features:
        if(feature == dependent_feature):
            continue
        if(not(test_data.dtypes[feature] in ['int64','float64'])):
            feature_list.append(feature)

    if(not(oot_data is None) and len(oot_data.index) > 0):
        oot_set = True
    if(not(pdv_data is None) and len(pdv_data.index) > 0):
        pdv_set = True

    cramersv_final_df = pd.DataFrame()
    if(pdv_set):
        if(oot_set):
            cv_temp_df = getChiSquare(oot_data,feature_list,dependent_feature,"OOT")
            #print(cv_temp_df)
            cv_temp_df.columns = ['Feature','Chi-Square_oot','P-Value_oot','CramersV_oot','Source']
            cv_temp_df.drop('Source',inplace=True,axis=1)
            cramersv_final_df = pd.concat([cramersv_final_df,cv_temp_df])            
            cramersv_final_df.reset_index(inplace=True)
            #cramersv_final_df.columns = ['Feature','CramersV_OOT'] 
            cv_temp_df = getChiSquare(pdv_data,feature_list,dependent_feature,"PDV")
            #print(cv_temp_df)
            temp_df = cv_temp_df.reset_index()
            temp_df.columns = ['index_pdv','Feature','Chi-Square_pdv','P-Value_pdv','CramersV_pdv','Source']
            temp_df.drop('Source',inplace=True,axis=1)
            temp_df.drop('index_pdv',inplace=True,axis=1)
            cramersv_final_df = cramersv_final_df.merge(temp_df,left_on='Feature',right_on='Feature')
            #print(cramersv_final_df)
        else:
            cv_temp_df = getChiSquare(pdv_data,feature_list,dependent_feature,"PDV")
            #print(cv_temp_df)
            cramersv_final_df = pd.concat([cramersv_final_df,cv_temp_df])
    else:
        if(oot_set):
            cv_temp_df = getChiSquare(oot_data,feature_list,dependent_feature,"OOT")
            cv_temp_df.columns = ['Feature','Chi-Square_oot','P-Value_oot','CramersV_oot','Source']
            cv_temp_df.drop('Source',inplace=True,axis=1)
            cramersv_final_df = pd.concat([cramersv_final_df,cv_temp_df])
        else:
            print('No OOT & PDV data provided hence skipping calculation')
            return
    cv_temp_df = getChiSquare(test_data,feature_list,dependent_feature,"Test")
    cv_temp_df.drop('Source',inplace=True,axis=1)
    cv_temp_df.columns = ['Feature','Chi-Square_test','P-Value_test','CramersV_test']
    cramersv_final_df = cramersv_final_df.merge(cv_temp_df,left_on='Feature',right_on='Feature')
    append_model_report(cramersv_final_df,'6.CramersV')
    
    #cramersv_final_df.to_csv(os.getcwd()+'/base/'+'temp/CramersV.csv')

def build_iv_report(final_df,features):
    sources = final_df['Source'].unique()
    iv_df = pd.DataFrame()
    if('Test' in sources):
        source_data = final_df[ final_df['Source'] == 'Test'  ]
        source_data = source_data[['Feature','IV']]
        source_data = source_data.groupby('Feature').sum()    
        iv_df = pd.concat([iv_df,source_data])
        iv_df.reset_index(inplace=True)
        iv_df.rename(columns={'IV':'IV_Test'},inplace=True)
        
    if('OOT' in sources):
        source_data = final_df[ final_df['Source'] == 'OOT'  ]
        source_data = source_data[['Feature','IV']]
        source_data = source_data.groupby('Feature').sum()
        source_data.reset_index(inplace=True)        
        iv_df = iv_df.merge(source_data,left_on="Feature",right_on="Feature")
        iv_df.rename(columns={'IV':'IV_OOT'},inplace=True)
        
    if('PDV' in sources):
        source_data = final_df[ final_df['Source'] == 'PDV'  ]
        source_data = source_data[['Feature','IV']]
        source_data = source_data.groupby('Feature').sum()
        source_data.reset_index(inplace=True)
        iv_df = iv_df.merge(source_data,left_on="Feature",right_on="Feature")        
        iv_df.rename(columns={'IV':'IV_PDV'},inplace=True)
    iv_df = iv_df.sort_values('IV_Test', ascending=False)
    append_model_report(iv_df,"5.IV")

def build_model_report(pipeline_obj,test,oot=None,pdv=None):

    book = load_workbook(os.getcwd()+'/base/'+'Reports/Post_Train_Model_Report.xlsx')
    writer = pd.ExcelWriter(os.getcwd()+'/base/'+'Reports/Post_Train_Model_Report.xlsx',engine='openpyxl')
    writer.book = book
    ws = book.create_sheet('1.Model Summary')
    ws.sheet_view.showGridLines = False
    #book['1.Model Summary']
    bold = Font(name='Calibri',size=14,bold=True)
    aln = Alignment(horizontal="left", vertical="center")
    double = Side(border_style="thin", color="000000")

    ws['A1'] = "Logistic Model Summary"
    ws['A1'].font = bold
    ws['A3'] = "Model Parameters & Coefficients"
    ws['A3'].font = bold

    items = len(pipeline_obj['model_features'])
    model_coefficients_count = len(pipeline_obj['model_coefficients'])
    
    for x in range(1,items+1):
        # if x == 1:
        #     ws.cell(row=5, column=x).value = pipeline_obj['final_feature_list'][x-1] 
        #     ws.cell(row=5, column=x).alignment = aln
        # else:
            ws.cell(row=5, column=x).value = pipeline_obj['model_features'][x-1]
            ws.cell(row=6, column=x).value = str(np.round(pipeline_obj['model_coefficients'][x-1],5))

            ws.cell(row=5, column=x).alignment = aln
            ws.cell(row=6, column=x).alignment = aln
    '''
    for x in range(1,len(pipeline_obj["model_coefficients"])+1):
       ws.cell(row=6, column=x).value = round(pipeline_obj['model_coefficients'][x-1],3)
       ws.cell(row=6, column=x).alignment = aln
    '''

    ws['A8'] = pipeline_obj['model_summary']
    ws['A8'].alignment = Alignment(horizontal='left',vertical='top',wrap_text=True)
    ws.merge_cells('A8:H'+str(22+model_coefficients_count))

    start_index_others = 20 + items + 4
    
    ws['A'+str(start_index_others)] = '--------------Testing Data Metrics--------------'
    ws['A'+str(start_index_others)].font = bold
    ws['A'+str(start_index_others+2)] = 'Confusion Matrix'
    ws['A'+str(start_index_others+2)].font = bold
    pipeline_obj['model_summary']
    ws['A'+str(start_index_others+4)] = 'Actual No'
    ws['B'+str(start_index_others+3)] = 'Predicted No'
    ws['A'+str(start_index_others+5)] = 'Actual Yes'
    ws['C'+str(start_index_others+3)] = 'Predicted Yes'    
    tn, fp, fn, tp = test["confusion_matrix"].ravel()
    ws['B'+str(start_index_others+4)] = tn
    ws['C'+str(start_index_others+4)] = fp
    ws['B'+str(start_index_others+5)] = fn
    ws['C'+str(start_index_others+5)] = tp

    ws['A'+str(start_index_others+7)] = 'ROC AUC Score'
    ws['A'+str(start_index_others+7)].font = bold
    ws['A'+str(start_index_others+8)] = round(test["roc_auc_score"],4)

    ws['A'+str(start_index_others+11)] = 'Classification Report'
    ws['A'+str(start_index_others+11)].font = bold
    ws['A'+str(start_index_others+13)] = str(test["classification_report"])
    ws['A'+str(start_index_others+13)].alignment = Alignment(horizontal='left',vertical='center',wrap_text=True)
    ws['A'+str(start_index_others+13)].border = Border(top=double, left=double, right=double, bottom=double)
    ws.merge_cells('A'+str(start_index_others+13)+":D"+str(start_index_others+23))

    if not((pdv is None)):
        ws['G'+str(start_index_others)] = '-------------- PDV  Metrics--------------'
        ws['G'+str(start_index_others)].font = bold
        ws['G'+str(start_index_others+2)] = 'Confusion Matrix'
        ws['G'+str(start_index_others+2)].font = bold
        
        ws['G'+str(start_index_others+4)] = 'Actual No'
        ws['H'+str(start_index_others+3)] = 'Predicted No'
        ws['G'+str(start_index_others+5)] = 'Actual Yes'
        ws['I'+str(start_index_others+3)] = 'Predicted Yes'    
        tn, fp, fn, tp = pdv["confusion_matrix"].ravel()
        ws['H'+str(start_index_others+4)] = tn
        ws['I'+str(start_index_others+4)] = fp
        ws['H'+str(start_index_others+5)] = fn
        ws['I'+str(start_index_others+5)] = tp

        ws['G'+str(start_index_others+7)] = 'ROC AUC Score'
        ws['G'+str(start_index_others+7)].font = bold
        ws['G'+str(start_index_others+8)] = round(pdv["roc_auc_score"],4)

        ws['G'+str(start_index_others+11)] = 'Classification Report'
        ws['G'+str(start_index_others+11)].font = bold
        ws['G'+str(start_index_others+13)] = str(pdv["classification_report"])
        ws['G'+str(start_index_others+13)].alignment = Alignment(horizontal='left',vertical='center',wrap_text=True)
        ws['G'+str(start_index_others+13)].border = Border(top=double, left=double, right=double, bottom=double)
        ws.merge_cells('G'+str(start_index_others+13)+':J'+str(start_index_others+23))
     
    if not((oot is None)):

        pdv_set = False
        if(not((pdv is None))):
            pdv_set = True

        if not(pdv_set):
            base_col = 'G'
            next_col = 'H'
            last_col = 'I'
            range_end_col = "J"
        else:
            base_col = 'M'
            next_col = 'N'
            last_col = 'O'
            range_end_col = "P"

        
        ws[base_col+str(start_index_others)] = '-------------- OOT  Metrics--------------'
        ws[base_col+str(start_index_others)].font = bold
        ws[base_col+str(start_index_others+2)] = 'Confusion Matrix'
        ws[base_col+str(start_index_others+2)].font = bold
        
        ws[base_col+str(start_index_others+4)] = 'Actual No'
        ws[next_col+str(start_index_others+3)] = 'Predicted No'
        ws[base_col+str(start_index_others+5)] = 'Actual Yes'
        ws[last_col+str(start_index_others+3)] = 'Predicted Yes'    
        tn, fp, fn, tp = oot["confusion_matrix"].ravel()
        ws[next_col+str(start_index_others+4)] = tn
        ws[last_col+str(start_index_others+4)] = fp
        ws[next_col+str(start_index_others+5)] = fn
        ws[last_col+str(start_index_others+5)] = tp

        ws[base_col+str(start_index_others+7)] = 'ROC AUC Score'
        ws[base_col+str(start_index_others+7)].font = bold
        ws[base_col+str(start_index_others+8)] = round(oot["roc_auc_score"],4)

        ws[base_col+str(start_index_others+11)] = 'Classification Report'
        ws[base_col+str(start_index_others+11)].font = bold
        ws[base_col+str(start_index_others+13)] = str(oot["classification_report"])

        ws[base_col+str(start_index_others+13)].border =  Border(top=double, left=double, right=double, bottom=double)
        ws[base_col+str(start_index_others+13)].alignment = Alignment(horizontal='left',vertical='center',wrap_text=True)
        ws.merge_cells(base_col+str(start_index_others+13)+':'+range_end_col+str(start_index_others+23))        
    writer.save()
    writer.close()

def compute_ks(dataset1, probability_feature_name, predicted_feature_name):
    dataset1 = dataset1.sort_values(probability_feature_name)
    dataset1 = dataset1.reset_index()
    target_count = dataset1[predicted_feature_name].sum()
    non_target_count = len(dataset1) - target_count
    dataset1['Cumulative_Population'] = dataset1.index + 1
    dataset1['Cumulative_Target'] = dataset1[predicted_feature_name].cumsum()
    dataset1['Cumulative_Non_Target'] = dataset1['Cumulative_Population'] - dataset1['Cumulative_Target']
    dataset1['Cumulative_Population_Percent'] = (dataset1['Cumulative_Population'] / dataset1.shape[0] )
    dataset1['Cumulative_Target_Percent'] =  dataset1['Cumulative_Target'] / target_count
    dataset1['Cumulative_Non_Target_Percent'] =  dataset1['Cumulative_Non_Target'] / non_target_count
    dataset1['ks'] = abs((dataset1['Cumulative_Target_Percent'] - dataset1['Cumulative_Non_Target_Percent']))    
    ks = dataset1['ks'].max()
    return ks

def compute_model_band_psi(model_risk_bands_train,model_risk_bands_test,model_risk_bands_pdv=None,model_risk_bands_oot=None):

    pdv_set = False
    oot_set = False

    model_band_psi_final_df = pd.DataFrame()

    total_population_train = model_risk_bands_train['Predicted'].count()
    total_population_test = model_risk_bands_test['Predicted'].count()
    

    if not(model_risk_bands_oot is None):
        total_population_oot = model_risk_bands_oot['Predicted'].count()        
        oot_set = True
    if not(model_risk_bands_pdv is None):
        total_population_pdv = model_risk_bands_pdv['Predicted'].count()        
        pdv_set = True

    
    psi_model_band_df = model_risk_bands_train.groupby(['Risk_Bands'], dropna=True).count().reset_index()       
    psi_model_band_df.columns = ['Risk_Bands_Train','Count_Train']    
    #total_population  = psi_model_band_df['Count_Train'].sum()
    
    band_count = psi_model_band_df['Risk_Bands_Train'].nunique()
    lift_per_band = 1.0 / band_count
    lifts =list(range(0,band_count))
    lifts[:] = [ lift_per_band * (i+1)*100.0 for i in lifts ]
    #lift_band = pd.Series(lifts,name='Lift')
    
    psi_model_band_df['Population_Train'] = psi_model_band_df['Count_Train'] / total_population_train       

    model_risk_bands_test.to_csv(os.getcwd()+'/base/'+'temp/model_risk_bands_test.csv')
    model_band_psi_final_df = pd.concat([model_band_psi_final_df,psi_model_band_df],axis=1)

    psi_model_band_df = model_risk_bands_test[['Risk_Bands','Actuals']].groupby(['Risk_Bands'], dropna=True).agg(['count','sum']).reset_index()       
    psi_model_band_df.columns = ['Risk_Bands_Test','Count_Test','Goods_Test']    
    psi_model_band_df['lift_per_band'] = pd.Series(lifts,name='Lift')
    #total_population  = psi_model_band_df['Count_Test'].sum()
    
    psi_model_band_df['Population_Test'] = psi_model_band_df['Count_Test'] / total_population_test  
    psi_model_band_df.to_csv(os.getcwd()+'/base/'+'temp/test-chk.csv')        
    psi_model_band_df['Bads_Test'] =  psi_model_band_df['Count_Test'] - psi_model_band_df['Goods_Test']
    psi_model_band_df['Cum_Bads_Test'] =  psi_model_band_df['Bads_Test'].cumsum()
    psi_model_band_df['Cum_Bads%_Test'] =  psi_model_band_df['Bads_Test'] / psi_model_band_df['Bads_Test'].sum()
    psi_model_band_df['Gain_Test'] =  round(psi_model_band_df['Cum_Bads%_Test'].cumsum() * 100.0, 4)
    psi_model_band_df['Lift_Test'] = psi_model_band_df['Gain_Test'] / psi_model_band_df['lift_per_band']

    model_band_psi_final_df = pd.concat([model_band_psi_final_df,psi_model_band_df],axis=1)

    if pdv_set:
        psi_model_band_df = model_risk_bands_pdv[['Risk_Bands','Actuals']].groupby(['Risk_Bands'], dropna=True).agg(['count','sum']).reset_index()       
        psi_model_band_df.to_csv(os.getcwd()+'/base/'+'temp/pdv-chk.csv') 
        psi_model_band_df.columns = ['Risk_Bands_PDV','Count_PDV','Goods_PDV']    
        #total_population  = psi_model_band_df['Count_PDV'].sum()
        psi_model_band_df['lift_per_band'] = pd.Series(lifts,name='Lift')
        psi_model_band_df['Population_PDV'] = psi_model_band_df['Count_PDV'] / total_population_pdv      
        psi_model_band_df['Bads_PDV'] =  psi_model_band_df['Count_PDV'] - psi_model_band_df['Goods_PDV']  
        psi_model_band_df['Cum_Bads_PDV'] =  psi_model_band_df['Bads_PDV'].cumsum()
        psi_model_band_df['Cum_Bads%_PDV'] =  psi_model_band_df['Bads_PDV'] / psi_model_band_df['Bads_PDV'].sum()
        psi_model_band_df['Gain_PDV'] =  round(psi_model_band_df['Cum_Bads%_PDV'].cumsum() * 100.0, 4)
        psi_model_band_df['Lift_PDV'] = psi_model_band_df['Gain_PDV'] / psi_model_band_df['lift_per_band']
        model_band_psi_final_df = pd.concat([model_band_psi_final_df,psi_model_band_df],axis=1)

    if oot_set:        
        psi_model_band_df = model_risk_bands_oot[['Risk_Bands','Actuals']].groupby(['Risk_Bands'], dropna=True).agg(['count','sum']).reset_index()       
        psi_model_band_df.to_csv(os.getcwd()+'/base/'+'temp/oot-chk.csv') 
        psi_model_band_df.columns = ['Risk_Bands_OOT','Count_OOT','Goods_OOT']    
        #total_population  = psi_model_band_df['Count_OOT'].sum()
        psi_model_band_df['lift_per_band'] = pd.Series(lifts,name='Lift')
        psi_model_band_df['Population_OOT'] = psi_model_band_df['Count_OOT'] / total_population_oot
        psi_model_band_df['Bads_OOT'] =  psi_model_band_df['Count_OOT'] - psi_model_band_df['Goods_OOT']
        psi_model_band_df['Cum_Bads_OOT'] =  psi_model_band_df['Bads_OOT'].cumsum()
        psi_model_band_df['Cum_Bads%_OOT'] =  psi_model_band_df['Bads_OOT'] / psi_model_band_df['Bads_OOT'].sum()
        psi_model_band_df['Gain_OOT'] =  round(psi_model_band_df['Cum_Bads%_OOT'].cumsum() * 100.0, 4)
        psi_model_band_df['Lift_OOT'] = psi_model_band_df['Gain_OOT'] / psi_model_band_df['lift_per_band']
        model_band_psi_final_df = pd.concat([model_band_psi_final_df,psi_model_band_df],axis=1)
    

    model_band_psi_final_df['PSI_Test'] = (model_band_psi_final_df['Population_Train'] - model_band_psi_final_df['Population_Test']) * np.log(model_band_psi_final_df['Population_Train']/model_band_psi_final_df['Population_Test'])
    model_band_psi_final_df['PSI_Test'] = model_band_psi_final_df['PSI_Test'].replace(to_replace=[np.inf,-np.inf],value=[0,0])
   
    if oot_set:
        model_band_psi_final_df['PSI_OOT'] = (model_band_psi_final_df['Population_Train'] - model_band_psi_final_df['Population_OOT']) * np.log(model_band_psi_final_df['Population_Train']/model_band_psi_final_df['Population_OOT'])
        model_band_psi_final_df['PSI_OOT'] = model_band_psi_final_df['PSI_OOT'].replace(to_replace=[np.inf,-np.inf],value=[0,0])
    
    if pdv_set:
        model_band_psi_final_df['PSI_PDV'] = (model_band_psi_final_df['Population_Train'] - model_band_psi_final_df['Population_PDV']) * np.log(model_band_psi_final_df['Population_Train']/model_band_psi_final_df['Population_PDV'])
        model_band_psi_final_df['PSI_PDV'] = model_band_psi_final_df['PSI_PDV'].replace(to_replace=[np.inf,-np.inf],value=[0,0])

    model_band_psi_final_df.to_csv(os.getcwd()+'/base/'+'temp/Model_Band_PSI.csv')       
    return model_band_psi_final_df

def plot_gains_lift(model_band_psi):
    book = load_workbook(os.getcwd()+'/base/'+'Reports/Post_Train_Model_Report.xlsx')
    writer = pd.ExcelWriter(os.getcwd()+'/base/'+'Reports/Post_Train_Model_Report.xlsx',engine='openpyxl')
    writer.book = book    
    #df.to_excel(writer,sheet_name=name,index=False)
    #print(book.sheetnames)
    gain_sheet =  book['9. ModelBand_PSI_Summary']
    
    start_index = model_band_psi.shape[0] + 5 

    test_col = 'A'
    oot_col = 'I'
    pdv_col = 'Q'    

    test_gain = model_band_psi['Gain_Test'].to_list()
    risk_bands = list(range(0,len(test_gain)))

    risk_bands_lab = risk_bands.copy()
    risk_bands_lab[:] = [ "Band "+str(risk_bands[x])  if x%2==0 else '' for x in risk_bands]

    #------------------------------------ Gain Table Test --------------------------------
    
    path = os.getcwd()+'/base/'+'temp/Gain'+'-'+'test'+'.png'
    plt.clf() 
    plt.xlabel('Risk Bands')
    plt.ylabel('Gain(%)')
    plt.xticks(ticks=risk_bands,labels=risk_bands_lab,rotation=45)
    plt.plot(risk_bands,test_gain)
    plt.savefig(path)

    img = openpyxl.drawing.image.Image(path)
    gain_sheet[test_col+str(start_index-1)] = 'Gain - Test'
    gain_sheet[test_col+str(start_index-1)].font = Font(name='Calibri',size=14,bold=True)
    gain_sheet.add_image(img, test_col+str(start_index))


    #------------------------------------ Lift Table Test --------------------------------
    test_lift = model_band_psi['Lift_Test'].to_list()
    lift_path = os.getcwd()+'/base/'+'temp/Lift'+'-'+'test'+'.png' 
    plt.clf() 
    plt.xlabel('Risk Bands')
    plt.ylabel('Lift')
    plt.xticks(ticks=risk_bands,labels=risk_bands_lab,rotation=45)
    plt.plot(risk_bands,test_lift)
    plt.savefig(lift_path)

    img = openpyxl.drawing.image.Image(lift_path)
    gain_sheet[test_col+str(start_index+34)] = 'Lift - Test'
    gain_sheet[test_col+str(start_index+34)].font = Font(name='Calibri',size=14,bold=True)
    gain_sheet.add_image(img, test_col+str(start_index+35))

    #--------------------------------------------------------------------

    
    cols = model_band_psi.columns 
    if 'Gain_OOT' in cols:
        oot_gain = model_band_psi['Gain_OOT'].to_list()
        risk_bands = list(range(0,len(test_gain)))        
        #model_band_psi['Risk_Bands_OOT'].astype(str).tolist()
        risk_bands_lab = risk_bands.copy()
        risk_bands_lab[:] = [ "Band "+str(risk_bands[x])  if x%2==0 else '' for x in risk_bands]          
        path = os.getcwd()+'/base/'+'temp/Gain'+'-'+'oot'+'.png'
        plt.clf() 
        plt.xlabel('Risk Bands')
        plt.ylabel('Gain(%)')
        plt.xticks(ticks=risk_bands,labels=risk_bands_lab,rotation=45) 
        plt.plot(risk_bands,oot_gain)
        plt.savefig(path)

        img = openpyxl.drawing.image.Image(path)
        gain_sheet[oot_col+str(start_index-1)] = 'Gain - OOT'
        gain_sheet[oot_col+str(start_index-1)].font = Font(name='Calibri',size=14,bold=True)
        gain_sheet.add_image(img, oot_col+str(start_index))

        oot_lift = model_band_psi['Lift_OOT'].to_list()
        lift_path = os.getcwd()+'/base/'+'temp/Lift'+'-'+'oot'+'.png'    
        plt.clf() 
        plt.xlabel('Risk Bands')
        plt.ylabel('Lift')
        plt.xticks(ticks=risk_bands,labels=risk_bands_lab,rotation=45)
        plt.plot(risk_bands,oot_lift)
        plt.savefig(lift_path)

        img = openpyxl.drawing.image.Image(lift_path)
        gain_sheet[oot_col+str(start_index+34)] = 'Lift - OOT'
        gain_sheet[oot_col+str(start_index+34)].font = Font(name='Calibri',size=14,bold=True)
        gain_sheet.add_image(img, oot_col+str(start_index+35))


    if  'Gain_PDV' in cols:
        pdv_gain = model_band_psi['Gain_PDV'].to_list()
        risk_bands = list(range(0,len(test_gain)))
        #model_band_psi['Risk_Bands_PDV'].astype(str).tolist()
        risk_bands_lab = risk_bands.copy()
        risk_bands_lab[:] = [ risk_bands[x]  if x%2==0 else '' for x in risk_bands]
        plt.xticks(ticks=risk_bands,labels=risk_bands_lab,rotation=45)
        path = os.getcwd()+'/base/'+'temp/Gain'+'-'+'pdv'+'.png'
        plt.clf() 
        plt.xlabel('Risk Bands')
        plt.ylabel('Gain(%)')        
        plt.plot(risk_bands,pdv_gain)
        plt.savefig(path)

        img = openpyxl.drawing.image.Image(path)
        gain_sheet[pdv_col+str(start_index-1)] = 'Gain - PDV'
        gain_sheet[pdv_col+str(start_index-1)].font = Font(name='Calibri',size=14,bold=True)
        gain_sheet.add_image(img, pdv_col+str(start_index))

        pdv_lift = model_band_psi['Lift_PDV'].to_list()
        lift_path = os.getcwd()+'/base/'+'temp/Lift'+'-'+'pdv'+'.png'    
        plt.clf() 
        plt.xlabel('Risk Bands')
        plt.ylabel('Lift')
        plt.xticks(ticks=risk_bands,labels=risk_bands_lab,rotation=45)
        plt.plot(risk_bands,pdv_lift)
        plt.savefig(lift_path)

        img = openpyxl.drawing.image.Image(lift_path)
        gain_sheet[pdv_col+str(start_index+34)] = 'Lift - PDV'
        gain_sheet[pdv_col+str(start_index+34)].font = Font(name='Calibri',size=14,bold=True)
        gain_sheet.add_image(img, pdv_col+str(start_index+35))

    writer.save()
    writer.close()


#---------------------------------------------------------------------------------------------------------

#Quick Run

#TODO add code specific to pdv in plot_all_bads
#obj = Reporting(None,None,None,None,None)
#obj.generate_reports()


#TODO List
# do away with pickle dependencies