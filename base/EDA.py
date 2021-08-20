import mlflow
from mlflow.tracking.fluent import log_params
import pandas as pd
import numpy as np
from scipy.stats import chi2_contingency
import traceback
import pickle
import json

import openpyxl
from statsmodels.stats.outliers_influence import variance_inflation_factor
from statsmodels.tools.tools import add_constant

from . import logistic_input as li
from sklearn.model_selection import train_test_split

#from custom_classes import *
#from model_refining import 
#from model_building import make_predictions, create_bin_model
from . import logistic_variable_model_selection
from .logistic_variable_model_selection import feature_selector_model
from .model_building import create_bin_model
from .model_refining import RefineBSModel
from .Reporting import * 

pd.options.display.max_columns = None
pd.options.display.max_rows = None
pd.options.display.precision = 2

import os
import yaml, io

from mlflow import log_metric, log_param, log_artifacts, log_metrics

from . import PipelineInputs as pl
from . import Fine_Tuning_Inputs as fti
from . import model_refining as ft


class ExecutionStepInputs:
    #Below are default Values
    DROP_MISSING_THRESHOLD = 0.6 
    BAD_OUTCOME = 'level1'
    IV_THRESHOLD = 0.02
    TARGET_VARIABLE = 'yTarget'
    TARGET_VARIABLE_GOOD_OUTCOME = 'Level0'
    TARGET_VARIABLE_BAD_OUTCOME = 'level1'
    CORRELATION_THRESHOLD = 0.5
    CRAMERSV_THRESHOLD = 0.5
    REPLACE_MISSING_WITH = 'mean'
    VIF_THRESHOLD = 10
    PSI_BUCKET_COUNT = 10
    CSI_THRESHOLD = 0.25
    EMPTY_CATEGORICAL_VALUE_REPLACEMENT = 'missing'
    EMPTY_NUMERIC_VALUE_REPLACEMENT = -999
    FEATURE_SELECTION_METHOD = 'RFE'
    MAX_SELECTED_FEATURES = 10
    RISK_BAND_COUNT = 5


html_template = "<!DOCTYPE html><html><head><title>Report</title><style type=\"text/css\">table,tr{border:1px solid black;border-spacing:0;text-align:center}tr th{background-color:#47a0ff;border-bottom:1px solid black;border-right:1px solid black;color:darkblue}tr td{border-right:1px solid black}</style></head><body>{table}</body></html>"


class LogisticPipeline:
    """docstring for LogisticPipeline"""
    def __init__(self, training_set_path,test_set_path,validate_set_path,ootd_set_path):
        super(LogisticPipeline, self).__init__()            
        self.log_steps = []
        self.training_set_path = training_set_path
        self.test_set_path = test_set_path
        self.validate_set_path = validate_set_path
        self.ootd_set_path = ootd_set_path
        self.final_feature_set = []
        self.dropped_features_set = []
        self.features_dict = {}
        self.pipeline_configuration = {}
        self.execution_steps = {}
        self.ivs = None
        self.csi_pd = pd.DataFrame()
        self.read_pipeline_configurations()
        self.percentile_bin_ranges = {}
        
        
        #self.create_metadata()

    def read_pipeline_configurations(self):
        try: 
            with open(os.getcwd()+'/base/'+pl.pipeline_input_file_path, 'r') as stream:
                model_configurations = yaml.safe_load(stream)                
                self.verify_configuration(model_configurations)
                self.pipeline_configuration = model_configurations
        except Exception as e:
            print('Error in Model Inputs file.')
            print(e)

    def verify_configuration(self,model_configurations):        
        # Checks if all mandatory keys are provided in YAML File
        for items in pl.pipeline_mandatory_keys:            
            if items not in model_configurations:
                print('Values for \'{}\' key mandatory to the Pipeline inputs file is not set'.format(items))
                raise Exception('Keys Mandatory for the Pipeline are {}'.format(pl.pipeline_mandatory_keys)) 
        self.execution_steps = model_configurations['set_execution_steps']
        #TODO Ideally Good to check data type of the Mandatory Keys as well to be sure

    def prepare_pipeline(self):
        #Create Logs,Model,Reports & temp diretory if not already present
        if not os.path.exists(os.getcwd()+'/base/'+self.pipeline_configuration['model_directory']):
            os.makedirs(os.getcwd()+'/base/'+self.pipeline_configuration['model_directory'])
        if not os.path.exists(os.getcwd()+'/base/'+self.pipeline_configuration['reports_directory']):
            os.makedirs(os.getcwd()+'/base/'+self.pipeline_configuration['reports_directory'])
            os.makedirs(os.getcwd()+'/base/'+self.pipeline_configuration['reports_directory']+'/PostTraining')
        if not os.path.exists(os.getcwd()+'/base/'+pl.DefaultInfo.default_staging_location):
            os.makedirs(os.getcwd()+'/base/'+pl.DefaultInfo.default_staging_location)
        if not os.path.exists(os.getcwd()+'/base/'+pl.DefaultInfo.default_staging_location+"/img"):
           os.makedirs(os.getcwd()+'/base/'+pl.DefaultInfo.default_staging_location+"/img")
        if os.path.exists(os.getcwd()+'/base/'+self.pipeline_configuration['reports_directory']+'/Pre_Train_Report.xlsx'):
            os.remove(os.getcwd()+'/base/'+self.pipeline_configuration['reports_directory']+'/Pre_Train_Report.xlsx')            
            wb = openpyxl.Workbook() 
            wb.save(filename= os.getcwd()+'/base/'+self.pipeline_configuration['reports_directory']+'/Pre_Train_Report.xlsx')
     
        # else:
        #     os.remove(pl.DefaultInfo.default_staging_location)
        #     os.makedirs(pl.DefaultInfo.default_staging_location)
        #Check if train,test,validate,oot dataset created for new cases
        #check

    def initialize_pipeline(self):
        
        data_directory = self.pipeline_configuration['data_directory']
                
        file_path = self.pipeline_configuration['training_dataset']
        print(file_path)
        if(file_path):
            self.training_set_path = file_path
            self.training_set = pd.read_csv(os.getcwd()+'/base/'+self.training_set_path,lineterminator='\n')
            self.log("Initialized training Dataframe")
        else:
            self.log("Training Set Path empty")
            raise Exception('Training Set Path is wrong/empty! This is Mandatory for the pipeline')

        file_path = self.pipeline_configuration['testing_dataset']
        print(file_path)
        if(file_path):
            self.test_set_path = file_path
            self.testing_set = pd.read_csv(os.getcwd()+'/base/'+self.test_set_path)
            self.log("Initialized testing Dataframe")
        else:
            self.log("Testing Set Path empty. Creating a testing by spliting the training")
            #self.training_set[TARGET_VARIABLE] = np.where(self.training_set[TARGET_VARIABLE].isin(['Level0']),1,0)            
            dtrain , dtest =  train_test_split(self.training_set, test_size=pl.DefaultInfo.default_train_test_split_ratio,random_state=144)
            dtrain.reset_index(inplace=True)
            dtest.reset_index(inplace=True)            
            self.training_set = dtrain
            self.testing_set = dtest
                        
        file_path = self.pipeline_configuration['validation_dataset']
        print(file_path)
        if(file_path):
            self.validate_set_path = file_path
            self.validate_set = pd.read_csv(os.getcwd()+'/base/'+self.validate_set_path)
            self.log("Initialized Validation Dataframe")
        else:
            self.validate_set = None
            self.log("Validation Set Path empty")

        file_path = self.pipeline_configuration['oot_dataset']
        print(file_path)
        if(file_path):
            self.ootd_set_path = file_path
            self.oot_set = pd.read_csv(os.getcwd()+'/base/'+self.ootd_set_path)
            self.log("Initialized OOT Dataframe")
        else:
            self.oot_set = None
            self.log("OOT Set Path empty")     
        

        transformed_data = li.input_data(projectname = self.pipeline_configuration['project_name'], 
            train = self.training_set, target = self.pipeline_configuration['target_variable'], 
            targettransformlist = self.pipeline_configuration['target_variable_bad_values'],
                 makebinomial=True, vars_exc=self.pipeline_configuration['excluded_variable_list'],
                 distcheck=False,write=self.pipeline_configuration['write_cleansed_data'],oos=self.testing_set,oot=self.oot_set,pdv=self.validate_set)
    
        self.training_set =  transformed_data['Train']        
        self.testing_set = transformed_data['OOS']
        self.oot_set = transformed_data['OOT']        
        self.validate_set = transformed_data['PDV']
        
        # self.training_set = self.training_set.sample(frac=0.1)      #Comment
        # self.training_set = self.training_set.reset_index()         #Comment
        # self.testing_set = self.testing_set.sample(frac=0.1)         #Comment    
        # self.testing_set = self.testing_set.reset_index()         #Comment
        # self.oot_set = self.oot_set.sample(frac=0.1)         #Comment    
        # self.oot_set = self.oot_set.reset_index()         #Comment
        
        self.final_feature_set = self.training_set.columns.tolist()
        self.setup_defaults()

    def setup_defaults(self):
        execution_step_inputs = self.pipeline_configuration['execution_step_inputs']
                
        ExecutionStepInputs.DROP_MISSING_THRESHOLD = execution_step_inputs['univariate_inputs']['drop_missing_value_threshold']
        ExecutionStepInputs.REPLACE_MISSING_WITH = execution_step_inputs['univariate_inputs']['impute_numeric_missing_values'] 
        ExecutionStepInputs.EMPTY_CATEGORICAL_VALUE_REPLACEMENT = execution_step_inputs['univariate_inputs']['impute_categorical_missing_values']
        ExecutionStepInputs.EMPTY_NUMERIC_VALUE_REPLACEMENT = -999
        
        ExecutionStepInputs.PSI_BUCKET_COUNT = execution_step_inputs['bivariate_inputs']['csi_bucket_counts']
        ExecutionStepInputs.CSI_THRESHOLD = execution_step_inputs['bivariate_inputs']['csi_threshold']
        ExecutionStepInputs.IV_THRESHOLD = execution_step_inputs['bivariate_inputs']['iv_threshold']
        ExecutionStepInputs.CRAMERSV_THRESHOLD = execution_step_inputs['bivariate_inputs']['cramersv_threshold']

        ExecutionStepInputs.VIF_THRESHOLD = execution_step_inputs['multivariate_inputs']['vif_threshold']

        ExecutionStepInputs.MAX_SELECTED_FEATURES = execution_step_inputs['feature_reduction']['max_shortlisted_features']
        ExecutionStepInputs.FEATURE_SELECTION_METHOD = execution_step_inputs['feature_reduction']['method']
        ExecutionStepInputs.RISK_BAND_COUNT = execution_step_inputs['model_building']['risk_bands']
        
    def create_metadata(self):
        self.features_dict['Numeric'] = []
        self.features_dict['Categorical'] = []

        for feature in self.training_set.columns:
            if self.training_set.dtype[feature] in ['object','category']:
                self.features_dict['Categorical'].append(feature)
            elif self.training_set.dtype[feature] in ['int64','float64']:
                self.features_dict['Numeric'].append(feature)
            elif self.training_set.dtype[feature] in ['bool']:
                self.features_dict['Categorical'].append(feature)
        #print(self.features_dict)

    def log(self,msg):
        self.log_steps.append(msg)
        self.log_steps.append('\n')

    def drop_missing_values(self):
        missing_data = self.training_set.isnull().sum()
        misssing_data_report = pd.DataFrame(missing_data,columns=['Missing Count'])
        total_rows = misssing_data_report['Missing Count'].max()
        misssing_data_report['Missing Percent'] = round((misssing_data_report['Missing Count'] / total_rows),3)
        drop_missing_items_list = misssing_data_report[misssing_data_report['Missing Percent'] >= ExecutionStepInputs.DROP_MISSING_THRESHOLD].index.tolist()        
        for feature in drop_missing_items_list:
            if  feature in self.final_feature_set:
                self.final_feature_set.remove(feature)
        #Eliminating Categorical Feature with cardinality > 15
        cardinality_list = []
        for feature in self.final_feature_set:
            if not(feature in self.training_set.select_dtypes(include=['number']).columns):
                if self.training_set[feature].nunique() > 15:
                    self.final_feature_set.remove(feature)
                    cardinality_list.append(feature)
        self.log("Dropping Below Features as they have more missing values than threshold {}".format(ExecutionStepInputs.DROP_MISSING_THRESHOLD))
        self.log(drop_missing_items_list)
        self.log("Eliminating Categorical Feature with cardinality > 15 {}".format(cardinality_list))
        #Exporting Dropped Features list to a csv
        #misssing_data_report.to_csv('Reports/Missing_Feature.csv',index=True)
        self.convert_df_to_html(misssing_data_report,'Reports','Missing_Feature',False)

    def replace_missing_value(self):        
        impute_steps = []
        self.log("Imputing Missing Values")
        impute_steps.append("Imputing Missing Values")

        handle_numeric_missing = ExecutionStepInputs.REPLACE_MISSING_WITH
        handle_categorical_missing = ExecutionStepInputs.EMPTY_CATEGORICAL_VALUE_REPLACEMENT

        for feature in self.final_feature_set:
            missing_value_count = self.training_set[feature].isnull().sum()
            
            if(missing_value_count == 0):
                continue #Nothing to do heres

            if(handle_numeric_missing == 'drop'):
                self.training_set[feature].dropna(inplace=True) #Dropping values
                continue

            if (self.training_set.dtypes[feature] in ["int64","float64"]):
                replace_value = 0
                if(handle_numeric_missing == 'mean'):
                    replace_value = self.training_set[feature].mean()
                if(handle_numeric_missing == 'median'):
                    replace_value = self.training_set[feature].median()
                if(handle_numeric_missing == 'constant'):
                    replace_value = -999

                self.training_set[feature].fillna(replace_value,inplace=True)
                if(missing_value_count > 0):
                    self.log("Replacing {} empty values in {} with {} = {}".format(missing_value_count,feature,ExecutionStepInputs.REPLACE_MISSING_WITH,replace_value))                
                    impute_steps.append("Replacing {} empty values in {} with {} = {}".format(missing_value_count,feature,ExecutionStepInputs.REPLACE_MISSING_WITH,replace_value))

            elif(self.training_set.dtypes[feature] in ["object"]): 
                if(missing_value_count == 0):               
                    self.training_set[feature].fillna(handle_categorical_missing,inplace=True)
                    self.log("Replacing {} empty values in {} with {}".format(self.training_set[feature].isnull().sum(),feature,handle_categorical_missing))                
                    impute_steps.append("Replacing {} empty values in {} with {}".format(self.training_set[feature].isnull().sum(),feature,handle_categorical_missing))
        
        write_output = pd.Series(impute_steps)
        write_output.name = 'Missing_Value_Imputations'
        self.convert_df_to_html(write_output.to_frame(),self.pipeline_configuration['reports_directory'],'Missing_Value_Imputations',hide_index=True)
        #write_output.to_csv("Reports/Missing_Variable_Replacement_Report.csv",index=False)

    def getCramerv(c,n,contigency):
        phi2 = c/n
        r,k = contigency.shape
        phi2corr = max(0.0, phi2 - (((k-1)*(r-1))/(n-1)))    
        rcorr = r - ((r-1)**2)/(n-1)
        kcorr = k - ((k-1)**2)/(n-1)
        cramers_v =  np.sqrt(phi2corr / min( (kcorr-1), (rcorr-1)))
        return cramers_v

    #Computing Chi-Square Statistics
    def getChiSquare(self, data, dependent_variable,):
        chi_sqaure_data = {'Feature_Name':[],'Chi-Square':[],'P-Value':[],'dof':[],'Significant':[],'CramersV':[]}
        alpha=0.05        
        categoricals = data[self.final_feature_set].select_dtypes(include=['object','category']).columns
        eliminated_features = []

        for feature in categoricals:
            if feature == dependent_variable:
                continue
            if(len(data[feature].unique()) > 15):   #Eliminating features with cardinality > 15
                eliminated_features.append(feature)
                continue

            contigency= pd.crosstab(data[dependent_variable],data[feature])
            c, p, dof, expected = chi2_contingency(contigency)
            n =  sum(contigency.sum())
            chi_sqaure_data["Feature_Name"].append(feature)
            chi_sqaure_data["Chi-Square"].append(c)
            chi_sqaure_data["P-Value"].append(p)
            chi_sqaure_data["dof"].append(dof)            
            chi_sqaure_data['CramersV'].append(getCramerv(c,n,contigency))
        chi_sqaure_pd = pd.DataFrame(chi_sqaure_data,columns=['Feature_Name','Chi-Square','P-Value','dof','CramersV'])
        
        self.convert_df_to_html(chi_sqaure_pd,"","CramersV")
        cramersv_filter = chi_sqaure_pd[chi_sqaure_pd["CramersV"]<=ExecutionStepInputs.CRAMERSV_THRESHOLD]["Feature_Name"].tolist()
        
        self.log('Eliminating following variables as CramersV below Threshold {}'.format(ExecutionStepInputs.CRAMERSV_THRESHOLD))  
        self.log(cramersv_filter)      
                
        self.log('Eliminating following variables as they have cardinality > 15 {}'.format(eliminated_features))
        self.log(eliminated_features)

        eliminated_features.extend(cramersv_filter)
        for feature in eliminated_features:
            self.final_feature_set.remove(feature)


    def get_variable_iv_score(self,feature_name,dataset, is_numeric=False):

        if (is_numeric):
            qlist = [np.round(i*0.1,1) for i in range(11)]
            quart = dataset[feature_name].quantile(qlist)
            temp_percetile_binning_train, feat_bins = pd.qcut(dataset[feature_name], 5 , precision=5, retbins=True, duplicates='drop' )
            self.percentile_bin_ranges[feature_name] = feat_bins            
        if is_numeric: 
            #check shape of the below 2
            contigency= pd.crosstab(temp_percetile_binning_train,dataset[ExecutionStepInputs.TARGET_VARIABLE],dropna=True)
        else:
            contigency= pd.crosstab(dataset[feature_name],dataset[ExecutionStepInputs.TARGET_VARIABLE],dropna=True)
        chi_sqaure_data = {'Feature_Name':[],'WOE':[],'IV':[]}        
        if (ExecutionStepInputs.TARGET_VARIABLE_GOOD_OUTCOME not in contigency.columns) and (ExecutionStepInputs.TARGET_VARIABLE_BAD_OUTCOME not in contigency.columns):
            print(feature_name)
            return 0
        #print(contigency)
        #contigency = contigency.T
        good_count = contigency[ExecutionStepInputs.TARGET_VARIABLE_GOOD_OUTCOME].sum()
        bad_count = contigency[ExecutionStepInputs.TARGET_VARIABLE_BAD_OUTCOME].sum()
        contigency["Percentage_Good"] = (contigency[ExecutionStepInputs.TARGET_VARIABLE_GOOD_OUTCOME]/good_count)
        contigency["Percentage_Bad"] = (contigency[ExecutionStepInputs.TARGET_VARIABLE_BAD_OUTCOME]/bad_count)
        contigency["woe"] = np.log(contigency["Percentage_Good"]/contigency["Percentage_Bad"])
        contigency["woe"] = contigency["woe"].replace([np.nan, np.inf, -np.inf],0)
        contigency["IV"] = (contigency["Percentage_Good"] - contigency["Percentage_Bad"]) * contigency["woe"]
        #print("{} IV {}".format(feature_name,contigency['IV'].sum()))
        return contigency['IV'].sum()

    def get_iv(self,feature):
        iv = self.ivs[feature]
        return iv

    def filter_by_iv(self,dataset):
        self.log("Computing IV scores for all independent variables - ")
        feature_iv = {"Feature":[],"IV":[]}

        numeric_features = dataset.select_dtypes(include=['number'])

        for feature in self.final_feature_set:
            feature_iv["Feature"].append(feature)
            iv = self.get_variable_iv_score(feature,dataset, feature in numeric_features)
            feature_iv["IV"].append(iv)   

        #iv_pd["IV"] = iv_pd["IV"].replace([np.nan, np.inf, -np.inf],0)
        iv_pd = pd.DataFrame(feature_iv)
        self.ivs = dict(iv_pd.values)
        iv_pd = iv_pd.sort_values('IV',ascending=False)

        
        iv_filter = iv_pd[iv_pd["IV"]<=ExecutionStepInputs.IV_THRESHOLD]["Feature"].tolist()
        #Filter out items based on iv < = 0.02 here & add to list.
        #print(iv_filter)
        #iv_pd.to_csv('Reports/IV_Report.csv',index=True)
        self.convert_df_to_html(iv_pd,self.pipeline_configuration['reports_directory'],'IV_Report',True)
        #self.log('{}'.format(iv_pd))
        self.log("Removing following features are they are below the IV Threshold of {}".format(ExecutionStepInputs.IV_THRESHOLD))        
        self.log(iv_filter)
        #self.convert_df_to_html(iv_filter,self.pipeline_configuration[''],'IV_Report')        
        for feature in iv_filter:
            self.final_feature_set.remove(feature)
                           
    def build_correlation_matrix(self,dataset):
        numeric_features = dataset[self.final_feature_set].select_dtypes(include=['number'])
        correlation_matrix = numeric_features.corr()
        correlation_matrix.index.name = 'numeric_features_list'

        correlation_matrix.reset_index(inplace=True)

        cols = correlation_matrix.columns
        eliminated_features = set()

        for feature1 in cols:
            for feature2 in cols:
                if (feature1 == feature2) or (feature1 == 'numeric_features_list') or (feature2 == 'numeric_features_list'):
                    continue
                correlation = correlation_matrix.loc[correlation_matrix['numeric_features_list'] == feature1 , feature2 ].values[0]
                if(correlation > ExecutionStepInputs.CORRELATION_THRESHOLD):
                    #Picking the feature with the higher IV of the 2 feature
                    if self.get_iv(feature1) > self.get_iv(feature2):
                        eliminated_features.add(feature2)
                    else:
                        eliminated_features.add(feature1)
                    
        self.log('Correlation > Threshold {} hence eleminating {}'.format(ExecutionStepInputs.CORRELATION_THRESHOLD,str(eliminated_features)))
        #print(eliminated_features)

        self.log("Generating Correlation Report")
        self.convert_df_to_html(correlation_matrix,self.pipeline_configuration['reports_directory'],'Correlation_Report')
        #self.log(correlation_matrix)
        
    def write_log(self):
        write_output = pd.Series(self.log_steps)
        write_output.to_csv(os.getcwd()+'/base/'+pl.DefaultInfo.default_log_path+"/Pipeline_Log.txt",index=False)

    def compute_vif(self,dataset):        
        feature_list = []
        for feature in self.final_feature_set:
            if(dataset.dtypes[feature] not in ['int64','float64']):
                continue
            else:
                feature_list.append(feature)        

        self.log("Computing VIFs to check for Multicolinearity between {} Features".format(feature_list))
        vif_ds = add_constant(dataset[feature_list])
        vifs = pd.Series([variance_inflation_factor(vif_ds.values, i) for i in range(vif_ds.shape[1])], index=vif_ds.columns)        
        #self.log(vifs)
        vifs = vifs.sort_values(ascending=False)
        self.convert_df_to_html(vifs.to_frame(),self.pipeline_configuration['reports_directory'],'Multicolinearity_Report',hide_index=False)
        eliminated_features = vifs[vifs >= ExecutionStepInputs.VIF_THRESHOLD].index.tolist()
        self.log("Removing features with VIFs greater than {}".format(ExecutionStepInputs.VIF_THRESHOLD))
        vifs = vifs[vifs <= ExecutionStepInputs.VIF_THRESHOLD]
        self.log("Eliminated following feature : {}".format(eliminated_features))
        #vifs.to_csv("Reports/Multicolinearity_Report.csv",index=True)
        
    def compute_psi(self,dataset1,dataset2):
        #get numeric & Cateogrical cols
        #numeric_features = list(dataset1.select_dtypes(include=['int64', 'float64']).columns)
        #categorical_features = list(dataset1.select_dtypes(include=['object']).columns)
        psi_dict = {"Feature":[],"PSI":[]}

        for feature in self.final_feature_set:
            print(f'{feature}')
            if (dataset1.dtypes[feature] not in ["int64","float64"]):
                continue
            self.log("Printing the feature - ".format(feature))           
            
            min_value = min(dataset1[feature].min(),dataset2[feature].min())
            print(f'Feature_Name = {feature}')
            print(f'min = {min_value}')
            max_value = max(dataset1[feature].max(),dataset2[feature].max())
            print(f'max = {max_value}')
            min_max_diff = round((max_value - min_value)/ExecutionStepInputs.PSI_BUCKET_COUNT)
            print(f'range = {min_max_diff}')
            bin_range = range(min_value,max_value,min_max_diff) 
            print(f'bin_range = {bin_range}')

            psi_df = pd.DataFrame({"dataset1":dataset1[feature],"dataset2":dataset2[feature]})
            psi_df["Expected"] = pd.cut(x=psi_df["dataset1"],bins=bin_range)
            psi_df["Actual"] = pd.cut(x=psi_df["dataset2"],bins=bin_range)
            
            psi_final = pd.concat([psi_df["Expected"].value_counts(),psi_df["Actual"].value_counts()],axis=1)
            expected_total = psi_final['Expected'].sum()
            actual_total = psi_final['Actual'].sum()
            psi_final['Expected'] =  psi_final['Expected']/expected_total 
            psi_final['Actual'] =  psi_final['Actual']/actual_total 
            psi_final['Difference'] = psi_final['Expected'] - psi_final['Actual']
            psi_final['log_division'] = np.log(psi_final['Expected'] / psi_final['Actual'])
            psi_final['psi'] = psi_final['Difference'] * psi_final['log_division']
            self.log(psi_final)
            psi = psi_final['psi'].sum()
            self.log("PSI = {}".format(psi))
            #psi_final.to_csv("psi.txt")
            psi_dict['Feature'].append(feature)
            psi_dict['PSI'].append(psi)
        
        psi_consolidated = pd.DataFrame(psi_dict)
        psi_consolidated.to_csv("PSI.txt")

    def compute_csi_numeric(self,series1,series2):
        #series1_endIndex = series1.count()

        min_value = round(min(series1.min(),series2.min()))         
        #print(f'min = {min_value}')
        max_value = round(max(series1.max(),series2.max()))
        #print(f'max = {max_value}')
        min_max_diff = round((max_value - min_value)/ExecutionStepInputs.PSI_BUCKET_COUNT)
        #print(f'range = {min_max_diff}')
        bin_range = [0.1,0.2,0.3,0.4,0.5,0.6,0.7,0.8,0.9,1.0]
        deciles = series1.quantile(bin_range)
        #print(f'bin_range = {bin_range}')

        df1 = pd.DataFrame(series1)
        df2 = pd.DataFrame(series2)    

        df1_bins = pd.Series(np.histogram(df1,bins=deciles)[0],name='expected')
        df2_bins = pd.Series(np.histogram(df2,bins=deciles)[0],name='actual')

        csi_df = pd.DataFrame([df1_bins,df2_bins])
        csi_df = csi_df.T
        
        #print(csi_df)        
        csi_df['expected'] = csi_df['expected'] / (csi_df['expected'].sum())
        csi_df['actual'] = csi_df['actual'] / (csi_df['actual'].sum())
        csi_df['diff'] = csi_df['actual'] - csi_df['expected']
        csi_df['psi'] = (csi_df['diff'])*np.log(csi_df['actual'] / csi_df['expected'])
        csi_df['psi'] = csi_df['psi'].replace([np.nan, np.inf, -np.inf],0)
        #print(csi_df)
        #print(csi_df['psi'].sum())
        #print(csi_df['psi'].sum())
        #print(csi_df['psi'])

        return csi_df['psi'].sum()
        #print(df.head(5))

    def compute_csi_categorical(self,series1,series2):
        
        df1_value_count = series1.value_counts()
        df2_value_count = series2.value_counts()

        df1_indexes = df1_value_count.index.tolist()
        df2_indexes = df2_value_count.index.tolist()

        categorical_values_list = list(set(df1_indexes+df2_indexes))        

        val = {"Values":categorical_values_list}
        df = pd.DataFrame(val)
        df['Expected'] = 0 
        df['Actual'] = 0 

        for categorical_value in categorical_values_list:
            if categorical_value in df1_indexes:
                df.loc[(df.Values == categorical_value),'Expected'] = df1_value_count.loc[categorical_value]
            if categorical_value in df2_indexes:
                df.loc[(df.Values == categorical_value),'Actual'] = df2_value_count.loc[categorical_value]

        df['Expected'] = df['Expected'] / df['Expected'].sum()
        df['Actual'] = df['Actual'] / df['Actual'].sum()

        df['Difference'] =  df['Actual'] - df['Expected']

        df['CSI'] =  (df['Difference']) * np.log(df['Actual'] / df['Expected']) 
        df['CSI'] = df['CSI'].replace([np.nan, np.inf, -np.inf],0)

        return df.CSI.sum()

    def compute_csi(self, dataset1 , dataset2 ):
        csi_dict = { 'Feature' : [], 'CSI' : [] }
        for feature in self.final_feature_set:
            if(feature in ["Unnamed: 0","emp_title","title"]):
                continue
            #print(feature)
            csi_dict['Feature'].append(feature)
            csi_value = 0

            if dataset1.dtypes[feature] in ['int64','float64']:
                csi_value = self.compute_csi_numeric(dataset1[feature],dataset2[feature])
            if dataset1.dtypes[feature] in ['object']:
                csi_value = self.compute_csi_categorical(dataset1[feature],dataset2[feature])                                        
            csi_dict['CSI'].append(round(csi_value,4))
        csi_df = pd.DataFrame(csi_dict) 
        #csi_df.to_csv("Reports/CSI.csv",index=False)
        self.convert_df_to_html(csi_df,self.pipeline_configuration['reports_directory'],'CSI_Report',hide_index=True)

    def kstable(train_pred, train_true, bad, bins=10, k_bin=None):
        score=pd.DataFrame(train_pred.values, columns=['SCORE']).reset_index (drop=True)
        ch=False
        if k_bin is None:
            ch=True
            k_bin=np.quantile(score['SCORE'],[np.round(i*0.1*10/bins,1) for i in range(bins+1)])
            k_bin[k_bin==min(k_bin)]=0
            k_bin[k_bin==max(k_bin)]=1
        
        k_bin.sort()
        score["DECILE"]=pd.cut(score['SCORE'],k_bin, include_lowest=True, precision=10, duplicates='drop').astype(str)
        score['TARGET']=train_true.values
        score['TARGET']=np.where(score['TARGET']==bad, 1,0)
        score['NONTARGET']= 1-score['TARGET']
        score=score.groupby(['DECILE']).sum().reset_index(drop=False)
        score['TOTAL']=score['TARGET']+score['NONTARGET']
        score=score.sort_values(["DECILE"], ascending=False).reset_index(drop=True)
        score['Population%']=np.round(score['TOTAL']/score['TOTAL'].sum()*100,2)
        score['BAD_CAPTURED' ]=np.round(score['TARGET']/score['TARGET'].sum()*100,2)
        score[ 'Cumulative_Target']=score['TARGET'].cumsum()
        score['Cumulative_NonTarget']=score['NONTARGET'].cumsum()
        score['Cumulative_Total']=score['TOTAL'].cumsum()
        score['Cumulative_%_Target/GainScore']=np.round(score['Cumulative_Target']/score['TARGET'].sum()*100,2)
        score['Cumulative_%_NonTarget']=np.round(score['Cumulative_NonTarget']/score['NONTARGET'].sum()*100,2)
        score['KS']=np.round(score['Cumulative_%_Target/GainScore' ]-score['Cumulative_%_NonTarget'],2)
        score['CumulativeLift']=score['Cumulative_%_Target/GainScore']/score['Population%']
        if ch:
            return score, k_bin
        else:
            return score

    def compute_ks(self,dataset1, probability_feature_name, predicted_feature_name):
        dataset1 = dataset1.sort_values(probability_feature_name)
        dataset1 = dataset1.reset_index()
        target_count = dataset1[predicted_feature_name].sum()
        non_target_count = len(dataset1) - target_count
        dataset1['Cumulative_Population'] = dataset1.index + 1
        dataset1['Cumulative_Target'] = dataset1[predicted_feature_name].cumsum()
        dataset1['Cumulative_Non_Target'] = dataset1['Cumulative_Population'] - dataset1['Cumulative_Target']
        dataset1['Cumulative_Population_Percent'] = (dataset1['Cumulative_Population'] / dataset1.shape[0] )
        dataset1['Cumulative_Target_Percent'] =  dataset1['Cumulative_Target'] / target_count
        dataset1['Cumulative_Non_Target_Percent'] =  dataset1['Cumulative_Non_Target'] / non_target_count
        dataset1['ks'] = abs((dataset1['Cumulative_Target_Percent'] - dataset1['Cumulative_Non_Target_Percent']))
        dataset1.to_csv("Reports/KS.csv")    
        ks = dataset1['ks'].max()
        #print(dataset1[predicted_feature_name,probability_feature_name,].head(1))
        #print(dataset1.tail(2))
        #print (dataset1.iloc[76900:77000,:])
        #plt.plot(dataset1[probability_feature_name],dataset1['Cumulative_Non_Target_Percent'],color='r')
        #plt.plot(dataset1[probability_feature_name],dataset1['Cumulative_Target_Percent'],color='b')
        return ks

    def log_mlflow_items(self):

        ml_flow_parameters = {}

        ml_flow_parameters["DROP_MISSING_THRESHOLD"] = ExecutionStepInputs.DROP_MISSING_THRESHOLD    
        ml_flow_parameters["BAD_OUTCOME"] = ExecutionStepInputs.BAD_OUTCOME    
        ml_flow_parameters["IV_THRESHOLD"] = ExecutionStepInputs.IV_THRESHOLD    
        ml_flow_parameters["TARGET_VARIABLE"] = ExecutionStepInputs.TARGET_VARIABLE    
        ml_flow_parameters["TARGET_VARIABLE_GOOD_OUTCOME"] = ExecutionStepInputs.TARGET_VARIABLE_GOOD_OUTCOME    
        ml_flow_parameters["TARGET_VARIABLE_BAD_OUTCOME"] = ExecutionStepInputs.TARGET_VARIABLE_BAD_OUTCOME    
        ml_flow_parameters["CORRELATION_THRESHOLD"] = ExecutionStepInputs.CORRELATION_THRESHOLD    
        ml_flow_parameters["REPLACE_MISSING_WITH"] = ExecutionStepInputs.REPLACE_MISSING_WITH    
        ml_flow_parameters["VIF_THRESHOLD"] = ExecutionStepInputs.VIF_THRESHOLD    
        ml_flow_parameters["PSI_BUCKET_COUNT"] = ExecutionStepInputs.PSI_BUCKET_COUNT    
        ml_flow_parameters["CSI_THRESHOLD"] = ExecutionStepInputs.CSI_THRESHOLD    
        ml_flow_parameters["EMPTY_CATEGORICAL_VALUE_REPLACEMENT"] = ExecutionStepInputs.EMPTY_CATEGORICAL_VALUE_REPLACEMENT    
        ml_flow_parameters["EMPTY_NUMERIC_VALUE_REPLACEMENT"] = ExecutionStepInputs.EMPTY_NUMERIC_VALUE_REPLACEMENT    
        ml_flow_parameters["FEATURE_SELECTION_METHOD"] = ExecutionStepInputs.FEATURE_SELECTION_METHOD

        param_json = json.dumps(ml_flow_parameters)

        # experiment = mlflow.get_experiment_by_name(self.pipeline_configuration['project_name'])

        # if(experiment is None):
        #     mlflow.create_experiment(self.pipeline_configuration['project_name'])

        log_param("Model Configurations", param_json)
        log_artifacts(self.pipeline_configuration['reports_directory'])
        log_artifacts(self.pipeline_configuration['model_directory'])
    
        self.log("Pipeline Execution complete with below configurations -  ")
        self.log(ml_flow_parameters)
        #build_correlation_matrix()      
        
    def describe_variables(self):
        deciles = [0.1,0.2,0.3,0.4,0.5,0.6,0.7,0.8,0.9]
        variable_desc_df = self.training_set.describe(percentiles=deciles,exclude=['object'])
        variable_desc_df = variable_desc_df.T
        variable_desc_df = variable_desc_df[['count','min','max','mean','std','10%','20%','30%','40%','50%','60%','70%','80%','90%']]        
        variable_desc_df = variable_desc_df.apply(pd.to_numeric)
        #variable_desc_df = np.round(variable_desc_df, 2)
        variable_desc_df = variable_desc_df.round(2)    
        self.convert_df_to_html(variable_desc_df,self.pipeline_configuration['reports_directory'],'Variable_Summary')

    def execute_pipeline(self):
        #Running code common to all steps        
        self.prepare_pipeline()


        #Quick Testing 
        #self.initialize_pipeline()  #comment
        #self.drop_missing_values()  #comment
        #self.replace_missing_value()    #comment    
        #self.getChiSquare(self.training_set, pl.DefaultInfo.default_converted_target_variable_name)           #comment

        if(self.execution_steps[pl.ExecutionStepsKey.univariate]):
            self.log('Begining Univariate Analysis')
            print('Begining Univariate Analysis')
            self.initialize_pipeline()
            self.describe_variables()
            self.drop_missing_values()
            self.replace_missing_value()            
            self.save_stage(self,pl.ExecutionStepsKey.univariate)
            print('End Univariate Analysis')
        if(self.execution_steps[pl.ExecutionStepsKey.bivariate]):       
            self.log('Begining Bivariate Analysis') 
            print('Begining Bivariate Analysis')                 
            self.filter_by_iv(self.training_set)
            self.build_correlation_matrix(self.training_set)
            self.getChiSquare(self.training_set, pl.DefaultInfo.default_converted_target_variable_name)
            self.save_stage(self,pl.ExecutionStepsKey.bivariate)
            print('End Bivariate Analysis')
        if(self.execution_steps[pl.ExecutionStepsKey.multivariate]):
            self.log('Begining Multivariate Analysis')
            print('Begining Multivariate Analysis')     
            self.compute_vif(self.training_set)
            self.save_stage(self,pl.ExecutionStepsKey.multivariate)
            print('End Multivariate Analysis')        
        if(self.execution_steps[pl.ExecutionStepsKey.feature_reduction]):
            self.log('Begining Feature Selection') 
            print('Begining Feature Selection') 
            #Stepwise / Lasso etc / Fw / Bk 
            shorlisted_features = self.final_feature_set
            shorlisted_features.append(ExecutionStepInputs.TARGET_VARIABLE)
            #self.training_set[ExecutionStepInputs.TARGET_VARIABLE] = np.where( self.training_set[ExecutionStepInputs.TARGET_VARIABLE] == 'Level0',0.0,1.0)
            self.training_set[ExecutionStepInputs.TARGET_VARIABLE] = self.training_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['Level0'],0.0)
            self.training_set[ExecutionStepInputs.TARGET_VARIABLE] = self.training_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['level1'],1.0)
            #self.training_set[ExecutionStepInputs.TARGET_VARIABLE] = self.training_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['Level0'],0.0)
            #self.training_set[ExecutionStepInputs.TARGET_VARIABLE] = self.training_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['level1'],1.0)
            #np.where(self.training_set[ExecutionStepInputs.TARGET_VARIABLE].isin(['Level0']),0.0,1.0)
            #self.testing_set[ExecutionStepInputs.TARGET_VARIABLE] = np.where(self.testing_set[ExecutionStepInputs.TARGET_VARIABLE].isin(['Level0']),0.0,1.0)

            self.testing_set[ExecutionStepInputs.TARGET_VARIABLE] = self.testing_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['Level0'],0.0)
            self.testing_set[ExecutionStepInputs.TARGET_VARIABLE] = self.testing_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['level1'],1.0)


            if(not(self.oot_set is None)):
                self.oot_set[ExecutionStepInputs.TARGET_VARIABLE] = self.oot_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['Level0'],0.0)
                self.oot_set[ExecutionStepInputs.TARGET_VARIABLE] = self.oot_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['level1'],1.0)

            if(not(self.validate_set is None)):
                self.validate_set[ExecutionStepInputs.TARGET_VARIABLE] = self.validate_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['Level0'],0.0)
                self.validate_set[ExecutionStepInputs.TARGET_VARIABLE] = self.validate_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['level1'],1.0)
            
            result_variables = feature_selector_model(data = self.training_set[shorlisted_features].copy(), 
                target = ExecutionStepInputs.TARGET_VARIABLE, numerical_feat_process = None, numerical_feat_missing = ExecutionStepInputs.EMPTY_NUMERIC_VALUE_REPLACEMENT, 
                        category_feat_process = 'woe', category_feat_missing= ExecutionStepInputs.EMPTY_CATEGORICAL_VALUE_REPLACEMENT, 
                        method = ExecutionStepInputs.FEATURE_SELECTION_METHOD) 

            feature_selection_result = result_variables['selected_features']
            self.log('Based on {} algorithm, the most significant features are {}'.format(ExecutionStepInputs.FEATURE_SELECTION_METHOD,feature_selection_result))
            self.final_feature_set = feature_selection_result
            '''if(not self.oot_set.empty):
                self.compute_csi(self.training_set[feature_selection_result],self.oot_set[feature_selection_result])
            else:
                self.log('Skipping CSI computation as OOT dataset not provided')'''
            self.save_stage(self,pl.ExecutionStepsKey.feature_reduction)
            print('End Feature Selection')
        if(self.execution_steps[pl.ExecutionStepsKey.model_building]):
            self.log('Begining Model Building')
            print('Begining Model Building')
            shorlisted_features = self.final_feature_set

            grid_param = { 
                'classifier__C' : np.logspace(-3, 1, 5),
                'classifier__penalty' : ['l1', 'l2', 'elasticnet', 'none'],
                'classifier__solver' : ['lbfgs','newton-cg','liblinear','sag','saga'],
                'classifier__max_iter' : [100, 1000, 2500, 5000]
            }
            
            result_model_pipelines = create_bin_model(data = self.training_set.copy(), target = ExecutionStepInputs.TARGET_VARIABLE, 
                variableset = self.final_feature_set, val_set = None, 
                get_Logit_summary = True ,penalty = 'l1', param_grid = grid_param,
                tune_hyperparams = False, risk_band_count = ExecutionStepInputs.RISK_BAND_COUNT)

            bins_info = self.get_bin_info(result_model_pipelines['model_bins'])
            risk_bands =  result_model_pipelines['pd_risk_bands']
            with open(pl.DefaultInfo.default_staging_location+"/model_preprocessor.pkl", 'wb') as file:
                pickle.dump(result_model_pipelines, file)

            with open(pl.DefaultInfo.default_staging_location+"/woe_bins.pkl", 'wb') as file:
                pickle.dump(bins_info, file)

            self.convert_df_to_html(bins_info,self.pipeline_configuration['reports_directory'],'Var.Bins',hide_index=True)
            self.convert_df_to_html(risk_bands,self.pipeline_configuration['reports_directory'],'Risk Bands')
            
            self.log('Result Model Pipeline - ')
            self.log(result_model_pipelines['model_pipeline'])

            firstcut_model = self.pipeline_configuration['model_directory']+"/first_cut_model.pkl"

            with open(firstcut_model, 'wb') as file:
                pickle.dump(result_model_pipelines['model_pipeline'], file)
            
            self.save_stage(self,pl.ExecutionStepsKey.model_building) 

            reports = Reporting(self.training_set,self.testing_set,self.validate_set,self.oot_set,None)
            reports.generate_reports()
            results = reports.get_mlflow_metrics()
            
            #self.log_mlflow_items()
            #mlflow.sklearn.log_model(result_model_pipelines,pl.DefaultInfo.default_model_path)            
            #mlflow.log_metric("KS",results['KS_Test']) 
            #mlflow.log_metric("AUC",results['AUC_Test'])
            print('End Model Building')
        if(self.execution_steps[pl.ExecutionStepsKey.fine_tuning]):
            #print(self.training_set[TARGET_VARIABLE].unique())
            self.log('Beginn Fine Tuning ')
            print('Beginn Fine Tuning ')
            pickle_file = pd.read_pickle(pl.DefaultInfo.default_staging_location+"/model_building.pkl")
            #self = pickle_file

            pickle_file.training_set[ExecutionStepInputs.TARGET_VARIABLE] = pickle_file.training_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['Level0'],0.0)
            pickle_file.training_set[ExecutionStepInputs.TARGET_VARIABLE] = pickle_file.training_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['level1'],1.0)  

            pickle_file.testing_set[ExecutionStepInputs.TARGET_VARIABLE] = pickle_file.testing_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['Level0'],0.0)
            pickle_file.testing_set[ExecutionStepInputs.TARGET_VARIABLE] = pickle_file.testing_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['level1'],1.0)


            #np.where(pickle_file.training_set[ExecutionStepInputs.TARGET_VARIABLE].isin(['Level0']),0.0,1.0)
            #pickle_file.testing_set[ExecutionStepInputs.TARGET_VARIABLE] = np.where(pickle_file.testing_set[ExecutionStepInputs.TARGET_VARIABLE].isin(['Level0']),0.0,1.0)
            if(not(pickle_file.oot_set is None)):
                pickle_file.oot_set[ExecutionStepInputs.TARGET_VARIABLE] = pickle_file.oot_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['Level0'],0.0)
                pickle_file.oot_set[ExecutionStepInputs.TARGET_VARIABLE] = pickle_file.oot_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['level1'],1.0)

            if(not(pickle_file.validate_set is None)):
                pickle_file.validate_set[ExecutionStepInputs.TARGET_VARIABLE] = pickle_file.validate_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['Level0'],0.0)
                pickle_file.validate_set[ExecutionStepInputs.TARGET_VARIABLE] = pickle_file.validate_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['level1'],1.0)

            #Reading pickle file before executing the model            
            
            
            if(pickle_file is None):
                raise ValueError('Model Building Pickle file not found in temp directory! Train Model first before fine tuning it!')

            #pickle_file.training_set[ExecutionStepInputs.TARGET_VARIABLE].iloc[60:100] = 1.0  #Comment Out
            #pickle_file.training_set[ExecutionStepInputs.TARGET_VARIABLE].iloc[8:50] = 0.0  #Comment Out
            
            result_model_pipelines = RefineBSModel(data =  pickle_file.training_set.copy() , 
                         target = pl.DefaultInfo.default_converted_target_variable_name,
                          variableset = fti.variableset , category_feat_missing = ExecutionStepInputs.EMPTY_CATEGORICAL_VALUE_REPLACEMENT, 
                           numeric_trasformations = fti.numeric_transformations , 
                            numeric_bining = fti.numeric_bining , custom_binning = fti.custom_bins , 
                           categorical_grouping = fti.categorical_grouping, 
                          get_Logit_summary = fti.get_Logit_summary, 
                         penalty = 'l1', tune_hyperparams = False, param_grid = None, get_summary_report = True, 
                            risk_band_count = ExecutionStepInputs.RISK_BAND_COUNT, custom_risk_bands = fti.custom_risk_bands)
                
            bins_info = self.get_bin_info(result_model_pipelines['model_bins'])
            
            with open(pl.DefaultInfo.default_staging_location+"/model_preprocessor.pkl", 'wb') as file:
                pickle.dump(result_model_pipelines, file)
    
            #predictions = RefineBSPredictions(test_data = pickle_file.testing_set.copy(), result_model_pipeline = refinedResults)
            
            self.log('Fine Tuning Result : Model Info')
            self.log(result_model_pipelines['model_pipeline'])

            firstcut_model = self.pipeline_configuration['model_directory']+"/first_cut_model.pkl"

            with open(firstcut_model, 'wb') as file:
                pickle.dump(result_model_pipelines['model_pipeline'], file)

            #TODO Test


            #self.save_stage(self,pl.ExecutionStepsKey.model_building)         
            reports = Reporting(pickle_file.training_set,pickle_file.testing_set,pickle_file.validate_set,pickle_file.oot_set,None,fine_tuneing=True)
            reports.generate_reports()
            results = reports.get_mlflow_metrics()

            #reports = Reporting(self.training_set,self.testing_set,self.validate_set,self.oot_set,None)
            #eports.generate_reports()
            

            #self.log_mlflow_items()
            mlflow.sklearn.log_model(result_model_pipelines,pl.DefaultInfo.default_model_path)
            mlflow.log_metric("KS",results['KS_Test']) 
            mlflow.log_metric("AUC",results['AUC_Test'])
            print('End Fine Tuning')

    def convert_df_to_html(self,dataframe, location, name, hide_index=False):                       
        html_output = html_template        
        table_contents = ""
        # if(hide_index):        
        #     table_contents = dataframe.style.hide_index().render()
        # else:
        #     table_contents = dataframe.style.render()
        # write html to file
        #html_output = html_template.replace('{table}', table_contents)

        #book = openpyxl.load_workbook(self.pipeline_configuration['reports_directory']+r'/Pre_Train_Report.xlsx')    
        #if not name in book.sheetnames:
        #    book.create_sheet(name)
        #book.save(self.pipeline_configuration['reports_directory']+r'/Pre_Train_Report.xlsx')

        if not os.path.exists(os.getcwd()+'/base/'+self.pipeline_configuration['reports_directory']+'/Pre_Train_Report.xlsx'):
            if not os.path.isfile(os.getcwd()+'/base/'+self.pipeline_configuration['reports_directory']+'/Pre_Train_Report.xlsx'):
                wb = openpyxl.Workbook() 
                cs = wb.active 
                cs.title = name
                wb.save(filename= os.getcwd()+'/base/'+self.pipeline_configuration['reports_directory']+'/Pre_Train_Report.xlsx')
        

        #dataframe.to_excel(self.pipeline_configuration['reports_directory']+r'/Pre_Train_Report.xlsx',sheet_name=name, engine='openpyxl')
        with pd.ExcelWriter(os.getcwd()+'/base/'+self.pipeline_configuration['reports_directory']+r'/Pre_Train_Report.xlsx', mode='a', engine='openpyxl',if_sheet_exists="replace") as writer:
            dataframe.to_excel(writer,sheet_name=name)
        #writer = pd.ExcelWriter(self.pipeline_configuration['reports_directory']+'/Pre_Train_Report.xlsx',engine='openpyxl')
        #writer.book = wb 
        #dataframe.to_excel(writer,sheet_name=name)        
        #writer.save()  
        #writer.close()      

        # report_html = open(location+'/'+name+'.html', "w")
        # report_html.write(html_output)
        # report_html.close()
    
    def test(self):

        self.initialize_dataframe()
        #self.drop_missing_values()
        #self.filter_by_iv(self.training_set)
        #self.replace_missing_value()
        #self.compute_vif(self.training_set)
        
        dataset2 = pd.read_csv("Data/loan_data_2015.csv")        
        #print(self.calculate_psi(self.training_set , dataset2, 'quantiles', buckets=10))

        #csi = self.compute_csi_numeric(self.training_set["annual_inc"],dataset2['annual_inc'])
        self.compute_csi(self.training_set,dataset2)
        self.write_log()
        '''
        self.filter_by_iv(self.training_set)
        
        correlation_matrix = self.training_set.corr()
        correlation_matrix_pairs = correlation_matrix.unstack(fill_value=1.0).reset_index()
        correlation_matrix_pairs.columns = ['Variable 1','Variable 2','Correlation']
        correlation_matrix_pairs['Absolute Corr'] = abs(correlation_matrix_pairs['Correlation'])

        correlation_matrix_pairs = correlation_matrix_pairs.sort_values(by='Absolute Corr', ascending=False)
        
        filtered_correlation_matrix_pairs = correlation_matrix_pairs[correlation_matrix_pairs['Absolute Corr'] < 1.0]
        filtered_correlation_matrix_pairs = filtered_correlation_matrix_pairs[filtered_correlation_matrix_pairs['Absolute Corr'] >= CORRELATION_THRESHOLD]

        correlation_filter_list = set()
        correlation_matrix = self.training_set.corr().abs()
        
        for i in range(len(correlation_matrix.columns)):
            for j in range(i):
        if(correlation_matrix.iloc[i,j] >= CORRELATION_THRESHOLD) and (correlation_matrix.columns[j] not in correlation_filter_list):
                    colname = correlation_matrix.columns[i]
                    correlation_filter_list.add(colname)
                    if colname in correlation_matrix.columns:
                        del correlation_matrix[colname]
        print(correlation_matrix)
        '''
    
    def clear_previous_stage(self,stage):
        previous_stage = ""
        if(stage == pl.ExecutionStepsKey.univariate):
            previous_stage = ""
        if(stage == pl.ExecutionStepsKey.bivariate):
            previous_stage = pl.ExecutionStepsKey.univariate
        if(stage == pl.ExecutionStepsKey.multivariate):
            previous_stage = pl.ExecutionStepsKey.bivariate
        if(stage == pl.ExecutionStepsKey.feature_reduction):
            previous_stage = pl.ExecutionStepsKey.multivariate
        if(stage == pl.ExecutionStepsKey.model_building):
            previous_stage = pl.ExecutionStepsKey.feature_reduction
        if(previous_stage == ""):
            return
        else:
            previous_stage_pkl = os.getcwd()+'/base/'+pl.DefaultInfo.default_staging_location+"/"+previous_stage+".pkl"
            if os.path.exists(previous_stage_pkl):
                os.remove(previous_stage_pkl)

    def save_stage(self,df,stage):
        self.clear_previous_stage(stage)
        pkl_filename = os.getcwd()+'/base/'+pl.DefaultInfo.default_staging_location+"/"+stage+".pkl"
        with open(pkl_filename, 'wb') as file:
            pickle.dump(df, file)

    def save_stage_kf(self,df,stage,container_output_path):
        pkl_filename = container_output_path +"/"+stage+".pkl"
        with open(pkl_filename, 'wb') as file:
            pickle.dump(df, file)
        #df.to_pickle(container_output_path)

    def get_bin_info(self,data):
        df = {}
        for feature in data.columns:
            if feature == 'Unnamed: 0':
                continue
            df[feature] =  data[feature].value_counts()
            df[feature].name = 'count'
            df[feature] = df[feature].reset_index()            
            df[feature]['feature'] = feature
            #print(df[feature]['index'].dtype)
            df[feature]['index'] = df[feature]['index'].astype('str')
            df[feature]['min'] = df[feature]['index'].apply(lambda x: x.split(',')[0][1:])        
        bins_info = pd.concat(df.values(),ignore_index=True)
        bins_info = bins_info[['feature','index','count','min']]     
        return bins_info

    def prepare_data(self,inputpath):
        self.log('Begining Univariate Analysis')
        print('Begining Univariate Analysis')
        self.initialize_pipeline()
        self.describe_variables()
        self.drop_missing_values()
        self.replace_missing_value()          
        self.save_stage(self,pl.ExecutionStepsKey.univariate)
        self.save_stage_kf(self,pl.ExecutionStepsKey.univariate,inputpath )
        print('End Univariate Analysis')

    def reduce_variables(self,input_path):
        #read and replace pickle & load others files for the stage here        
        if os.path.isfile(input_path+'/univariate.pkl'):
            pkl = pd.read_pickle(input_path+'/univariate.pkl')
            os.remove(input_path+'/univariate.pkl')
        else:
            raise ValueError('Previous Stage Pickle file not found')
        self = pkl
        #delete old pkl
        #call next set of functions
        self.log('Begining Bivariate & Multivariate Analysis') 
        print('Begining Bivariate & Multivariate Analysis')                 
        self.filter_by_iv(self.training_set)
        self.build_correlation_matrix(self.training_set)
        self.getChiSquare(self.training_set, pl.DefaultInfo.default_converted_target_variable_name)
        self.compute_vif(self.training_set)
        self.save_stage(self,pl.ExecutionStepsKey.bivariate)
        self.save_stage_kf(self,pl.ExecutionStepsKey.bivariate, input_path)
        print('End Bivariate Analysis')

    def feature_selection(self,input_path):
        if os.path.isfile(input_path+'/bivariate.pkl'):
            pkl = pd.read_pickle(input_path+'/bivariate.pkl')
            os.remove(input_path+'/bivariate.pkl')
        else:
            raise ValueError('Previous Stage Pickle file not found')
        self = pkl
        #call next set of functions
        self.log('Begining Feature Selection') 
        print('Begining Feature Selection') 
        shorlisted_features = self.final_feature_set
        #Stepwise / Lasso etc / Fw / Bk 
        shorlisted_features = self.final_feature_set
        shorlisted_features.append(ExecutionStepInputs.TARGET_VARIABLE)
        self.training_set[ExecutionStepInputs.TARGET_VARIABLE] = self.training_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['Level0'],0.0)
        self.training_set[ExecutionStepInputs.TARGET_VARIABLE] = self.training_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['level1'],1.0)
        self.testing_set[ExecutionStepInputs.TARGET_VARIABLE] = self.testing_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['Level0'],0.0)
        self.testing_set[ExecutionStepInputs.TARGET_VARIABLE] = self.testing_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['level1'],1.0)

        if(not(self.oot_set is None)):
            self.oot_set[ExecutionStepInputs.TARGET_VARIABLE] = self.oot_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['Level0'],0.0)
            self.oot_set[ExecutionStepInputs.TARGET_VARIABLE] = self.oot_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['level1'],1.0)

        if(not(self.validate_set is None)):
            self.validate_set[ExecutionStepInputs.TARGET_VARIABLE] = self.validate_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['Level0'],0.0)
            self.validate_set[ExecutionStepInputs.TARGET_VARIABLE] = self.validate_set[ExecutionStepInputs.TARGET_VARIABLE].replace(['level1'],1.0)
            
        result_variables = feature_selector_model(data = self.training_set[shorlisted_features].copy(), 
            target = ExecutionStepInputs.TARGET_VARIABLE, numerical_feat_process = None, numerical_feat_missing = ExecutionStepInputs.EMPTY_NUMERIC_VALUE_REPLACEMENT, 
                    category_feat_process = 'woe', category_feat_missing= ExecutionStepInputs.EMPTY_CATEGORICAL_VALUE_REPLACEMENT, 
                    method = ExecutionStepInputs.FEATURE_SELECTION_METHOD) 

        feature_selection_result = result_variables['selected_features']
        self.log('Based on {} algorithm, the most significant features are {}'.format(ExecutionStepInputs.FEATURE_SELECTION_METHOD,feature_selection_result))
        self.final_feature_set = feature_selection_result
        '''if(not self.oot_set.empty):
            self.compute_csi(self.training_set[feature_selection_result],self.oot_set[feature_selection_result])
        else:
            self.log('Skipping CSI computation as OOT dataset not provided')'''
        self.save_stage(self,pl.ExecutionStepsKey.feature_reduction)
        self.save_stage_kf(self,pl.ExecutionStepsKey.feature_reduction, input_path)
        print('End Feature Selection') 
          
    def model_building(self,input_path):
        pkl = pd.read_pickle(input_path+'/feature_reduction.pkl')
        self = pkl
        
        self.log('Begining Model Building')
        print('Begining Model Building')
        
        shorlisted_features = self.final_feature_set

        grid_param = { 
            'classifier__C' : np.logspace(-3, 1, 5),
            'classifier__penalty' : ['l1', 'l2', 'elasticnet', 'none'],
            'classifier__solver' : ['lbfgs','newton-cg','liblinear','sag','saga'],
            'classifier__max_iter' : [100, 1000, 2500, 5000]
        }
        
        result_model_pipelines = create_bin_model(data = self.training_set.copy(), target = ExecutionStepInputs.TARGET_VARIABLE, 
            variableset = self.final_feature_set, val_set = None, 
            get_Logit_summary = True ,penalty = 'l1', param_grid = grid_param,
            tune_hyperparams = False, risk_band_count = ExecutionStepInputs.RISK_BAND_COUNT)
        
        bins_info = self.get_bin_info(result_model_pipelines['model_bins'])
        risk_bands =  result_model_pipelines['pd_risk_bands']
        with open(os.getcwd()+'/base/'+pl.DefaultInfo.default_staging_location+"/model_preprocessor.pkl", 'wb') as file:
            pickle.dump(result_model_pipelines, file)

        with open(os.getcwd()+'/base/'+pl.DefaultInfo.default_staging_location+"/woe_bins.pkl", 'wb') as file:
            pickle.dump(bins_info, file)

        #firstcut_model = self.pipeline_configuration['model_directory']+"/first_cut_model.pkl"

        self.convert_df_to_html(bins_info,self.pipeline_configuration['reports_directory'],'Var.Bins',hide_index=True)
        self.convert_df_to_html(risk_bands,self.pipeline_configuration['reports_directory'],'Risk Bands')
            
        self.log('Result Model Pipeline - ')
        self.log(result_model_pipelines['model_pipeline'])

        firstcut_model = os.getcwd()+'/base/'+self.pipeline_configuration['model_directory']+"/first_cut_model.pkl"

        kf_loc = input_path + "/model_preprocessor.pkl"
        with open(kf_loc, 'wb') as file:
            pickle.dump(result_model_pipelines, file)

        kf_loc = input_path + "/first_cut_model.pkl"
        with open(kf_loc, 'wb') as file:
            pickle.dump(result_model_pipelines['model_pipeline'], file)

        self.save_stage(self,pl.ExecutionStepsKey.model_building)
        self.save_stage_kf(self,pl.ExecutionStepsKey.model_building,input_path)
 
        reports = Reporting(self.training_set,self.testing_set,self.validate_set,self.oot_set,None)
        reports.generate_reports()

        results = reports.get_mlflow_metrics()
        
        import json

        with open(input_path+'/results.json', 'wb') as fp:
            json.dump(results, fp)

        with open(input_path+'/config.json', 'wb') as fp:
            json.dump(self.pipeline_configuration, fp)
    
        # mlflow.sklearn.log_model(result_model_pipelines,pl.DefaultInfo.default_model_path)           
        # mlflow.log_metric("KS",results['KS_Test']) 
        # mlflow.log_metric("AUC",results['AUC_Test'])


        print('End Model Building')    

if __name__=='__main__':
    LogisticPipeline_obj = LogisticPipeline("","","","")    
    try:
        LogisticPipeline_obj.log('Start')        
        LogisticPipeline_obj.execute_pipeline()
        LogisticPipeline_obj.log('End')
    except Exception as e:
        LogisticPipeline_obj.log(e)
        LogisticPipeline_obj.log(traceback.format_exc())
        
    finally:
        LogisticPipeline_obj.write_log()

def exeute():
    LogisticPipeline_obj = LogisticPipeline("","","","")    
    try:
        LogisticPipeline_obj.log('Start')        
        LogisticPipeline_obj.execute_pipeline()
        LogisticPipeline_obj.log('End')
    except Exception as e:
        LogisticPipeline_obj.log(e)
        LogisticPipeline_obj.log(traceback.format_exc())
        
    finally:
        LogisticPipeline_obj.write_log()





#TODO -- 
# 1. Create a list of Numeric & Categorical Features in the constructor of the Class
# 2. replace all repalce with mean with simple imputer & standerd scaler
# 3. Refactor the CSI code to run PSI.
# 4. Define Global method get_default_pipeline_parameter() to run pipeline as a module